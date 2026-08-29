# -*- coding: utf-8 -*-
import os
import json
import csv
import math
import uuid
import zlib
from datetime import datetime

from kivy.config import Config
from kivy.utils import escape_markup, platform

# =========================
# 字体设置
# =========================
# 电脑端：优先用系统安装的 Noto Serif SC
# 安卓端：用项目目录里的 NotoSerifSC-Regular.otf
APP_DIR = os.path.dirname(os.path.abspath(__file__))
WINDOWS_FONT_PATH = r"C:\Windows\Fonts\NotoSerifSC-Regular.ttf"
LOCAL_FONT_PATH = os.path.join(APP_DIR, "NotoSerifSC-Regular.otf")

font_path = None
if platform == "win" and os.path.exists(WINDOWS_FONT_PATH):
    font_path = WINDOWS_FONT_PATH
elif os.path.exists(LOCAL_FONT_PATH):
    font_path = LOCAL_FONT_PATH

if font_path:
    Config.set(
        "kivy",
        "default_font",
        ["AppFont", font_path, font_path, font_path, font_path]
    )

from kivy.app import App
from kivy.core.window import Window
from kivy.metrics import dp, sp
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.behaviors import ButtonBehavior
from kivy.uix.button import Button
from kivy.uix.filechooser import FileChooserListView
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.popup import Popup
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.uix.scrollview import ScrollView
from kivy.uix.spinner import Spinner, SpinnerOption
from kivy.uix.switch import Switch
from kivy.uix.textinput import TextInput
from kivy.uix.widget import Widget
from kivy.graphics import Color, Ellipse, Line, RoundedRectangle
from kivy.clock import Clock
from kivy.animation import Animation
from kivy.properties import NumericProperty

from openpyxl import Workbook, load_workbook

# 桌面端设置最小窗口，安卓端不受影响
if platform in ("win", "linux", "macosx"):
    Window.minimum_width = 380
    Window.minimum_height = 700

# =========================
# 移动端视觉主题
# =========================
COLOR_PAGE_BG = (0.969, 0.973, 0.980, 1)       # F7F8FA
COLOR_CARD_BG = (1, 1, 1, 1)                   # FFFFFF
COLOR_PRIMARY = (0.145, 0.388, 0.922, 1)       # 2563EB
COLOR_PRIMARY_LIGHT = (0.937, 0.965, 1, 1)      # EFF6FF
COLOR_TEXT = (0.067, 0.094, 0.153, 1)           # 111827
COLOR_TEXT_SECONDARY = (0.420, 0.447, 0.502, 1) # 6B7280
COLOR_BORDER = (0.898, 0.906, 0.922, 1)         # E5E7EB
COLOR_DANGER = (0.863, 0.149, 0.149, 1)         # DC2626
COLOR_DANGER_LIGHT = (0.996, 0.949, 0.949, 1)   # FEF2F2
COLOR_SUCCESS = (0.086, 0.639, 0.290, 1)        # 16A34A
COLOR_SUCCESS_LIGHT = (0.941, 0.992, 0.957, 1)  # F0FDF4
COLOR_WHITE = (1, 1, 1, 1)
COLOR_TRANSPARENT = (1, 1, 1, 0)

PAGE_PADDING = dp(16)
CARD_SPACING = dp(12)
CARD_PADDING = dp(16)
CARD_RADIUS = dp(16)
BUTTON_RADIUS = dp(12)
INPUT_RADIUS = dp(10)
PRIMARY_BUTTON_HEIGHT = dp(52)
CONTROL_HEIGHT = dp(48)


CATEGORY_CHART_COLORS = [
    (0.91, 0.30, 0.24, 1),
    (0.20, 0.60, 0.86, 1),
    (0.18, 0.80, 0.44, 1),
    (0.95, 0.61, 0.07, 1),
    (0.61, 0.35, 0.71, 1),
    (0.10, 0.74, 0.61, 1),
    (0.90, 0.49, 0.13, 1),
    (0.20, 0.29, 0.37, 1),
    (0.84, 0.15, 0.45, 1),
    (0.40, 0.70, 0.20, 1),
]


# 分类设置项整行背景的极浅柔和调色板：黑字可读、操作按钮（蓝/红/绿）仍可区分
CATEGORY_BG_COLORS = [
    (0.93, 0.95, 0.98, 1),   # 极淡蓝
    (0.94, 0.97, 0.93, 1),   # 极淡绿
    (0.98, 0.96, 0.91, 1),   # 极淡橙
    (0.97, 0.93, 0.96, 1),   # 极淡粉
    (0.94, 0.95, 0.98, 1),   # 极淡紫
    (0.93, 0.97, 0.96, 1),   # 极淡青
    (0.98, 0.94, 0.93, 1),   # 极淡珊瑚
    (0.96, 0.96, 0.97, 1),   # 极淡灰
]


def _category_bg_color(category):
    """确定性颜色映射：按分类名字节做 crc32 取模。

    不使用内置 hash()（Python 进程内随机盐导致跨进程/重启不一致）；
    crc32 只依赖分类名本身，同一分类在任何刷新/重启/排序下颜色恒定。
    """
    idx = zlib.crc32(category.encode("utf-8")) % len(CATEGORY_BG_COLORS)
    return CATEGORY_BG_COLORS[idx]


# 默认分类：仅在 categories.json 缺失、损坏或没有任何有效分类时作为回退
DEFAULT_CATEGORIES = ["饮食正餐", "娱乐消费", "学习提升", "交通", "水电", "人情世故", "房租", "医疗", "其他"]


# 底部页面顺序（左→右），用于切换动画方向；settings 不参与排序
# income 仅在“启用收入功能”开启时出现在底栏，关闭时不可达
PAGE_ORDER = {
    "accounting": 0,
    "records": 1,
    "income": 2,
    "stats": 3,
}


class ThemedSpinnerOption(SpinnerOption):
    """保证下拉选项在 Android 上保持清晰且具备足够触控高度。"""
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.size_hint_y = None
        self.height = CONTROL_HEIGHT
        self.background_normal = ""
        self.background_disabled_normal = ""
        self.background_color = COLOR_TRANSPARENT
        self.color = COLOR_PRIMARY
        self.font_size = sp(17)

        with self.canvas.before:
            Color(*COLOR_PRIMARY_LIGHT)
            self.option_bg = RoundedRectangle(pos=self.pos, size=self.size)

        def update_background(instance, value):
            instance.option_bg.pos = instance.pos
            instance.option_bg.size = instance.size

        self.bind(pos=update_background, size=update_background)


class RecordRow(ButtonBehavior, BoxLayout):
    """统一的“单笔账目记录”卡片（支出/收入/分类消费详情/单笔消费排名共用）。

    结构（两行标准 + 可选序号行）：
      可选：序号（rank 非 None 时单独一行，左对齐）
      第一行：日期 · 分类（整卡宽度，单行稳定：正常字号 → 轻微缩小 → 单行省略）
      第二行：备注（弹性，按真实宽度适配） + 金额（固定宽度，右对齐）

    排名场景传入 rank=1..N，其余场景 rank=None（不显示序号行）。
    """
    def __init__(self, record, on_open, rank=None, **kwargs):
        super().__init__(
            orientation="vertical", spacing=dp(4), padding=[dp(14), dp(10)],
            size_hint_y=None, height=dp(100 if rank is not None else 76), **kwargs
        )
        self.record = record
        with self.canvas.before:
            Color(*COLOR_BORDER)
            self.row_border = RoundedRectangle(pos=self.pos, size=self.size, radius=[BUTTON_RADIUS] * 4)
            Color(*COLOR_CARD_BG)
            self.row_bg = RoundedRectangle(
                pos=(self.x + dp(1), self.y + dp(1)),
                size=(max(0, self.width - dp(2)), max(0, self.height - dp(2))),
                radius=[BUTTON_RADIUS - dp(1)] * 4
            )

        def update_canvas(instance, value):
            instance.row_border.pos = instance.pos
            instance.row_border.size = instance.size
            instance.row_bg.pos = (instance.x + dp(1), instance.y + dp(1))
            instance.row_bg.size = (max(0, instance.width - dp(2)), max(0, instance.height - dp(2)))

        self.bind(pos=update_canvas, size=update_canvas)
        self.bind(on_release=lambda instance: on_open(self.record))

        # ---- 可选：序号行（排名专用，单独一行，左对齐，不占用正文宽度） ----
        if rank is not None:
            rank_label = Label(
                text=str(rank), color=COLOR_PRIMARY, font_size=sp(18), bold=True,
                halign="left", valign="middle", size_hint_y=None, height=dp(22)
            )
            rank_label.bind(size=lambda inst, val: setattr(inst, "text_size", (val[0], None)))
            self.add_widget(rank_label)

        # ---- 第一行：日期 · 分类（整卡宽度、左对齐、单行稳定） ----
        meta = Label(
            text=f"{record.get('日期', '')}  ·  {record.get('分类', '')}",
            color=COLOR_TEXT_SECONDARY, font_size=sp(13), halign="left", valign="middle",
            size_hint_y=None, height=dp(20), shorten=True, shorten_from="right", max_lines=1
        )
        _fit_single_line_label(meta, sp(11), sp(13))
        self.add_widget(meta)

        # ---- 第二行：备注（按真实剩余宽度适配） + 金额（固定宽度，右对齐） ----
        content_row = BoxLayout(
            orientation="horizontal", spacing=dp(12),
            size_hint_y=None, height=dp(26)
        )
        note = Label(
            text=str(record.get("姓名/备注", "")), color=COLOR_TEXT, font_size=sp(16),
            halign="left", valign="middle", shorten=True, shorten_from="right", max_lines=1,
            size_hint_y=None, height=dp(26)
        )
        _fit_single_line_label(note, sp(13), sp(16))
        content_row.add_widget(note)

        try:
            amount_text = f"{float(record.get('金额', 0)):.2f} 元"
        except (TypeError, ValueError):
            amount_text = f"{record.get('金额', '')} 元"
        amount = Label(
            text=amount_text, color=COLOR_PRIMARY, font_size=sp(17), bold=True,
            halign="right", valign="middle", size_hint_x=None, width=dp(120),
            size_hint_y=None, height=dp(26), shorten=False, max_lines=1
        )
        amount.bind(size=lambda inst, val: setattr(inst, "text_size", (val[0], None)))
        content_row.add_widget(amount)
        self.add_widget(content_row)


def _fit_single_line_label(label, min_font, max_font):
    """单行三级适配：正常字号单行 → 轻微缩小字号 → 单行省略（延迟收敛）。

    绑定 label 的 size 变化；布局宽度稳定后按真实 texture 宽度决策，
    避免用布局前的临时宽度误缩字号，并防止缩小/恢复振荡。
    """
    label._fit_lock = False
    label._fit_evt = None
    label._last_shrink = False

    def decide(inst, avail):
        inst._fit_lock = True
        try:
            inst.text_size = (None, None)  # 完整渲染，测量真实宽度
            Clock.schedule_once(lambda dt: settle(inst, avail), 0.05)
        finally:
            inst._fit_lock = False

    def settle(inst, avail):
        if inst._fit_lock:
            Clock.schedule_once(lambda dt: settle(inst, avail), 0.05)
            return
        inst._fit_lock = True
        try:
            inst.texture_update()  # 强制按当前字号重建纹理，否则 texture_size 返回旧值
            tex_w = inst.texture_size[0]
            if tex_w > avail:
                if inst.font_size > min_font:
                    inst.font_size = max(min_font, inst.font_size - sp(1))
                    inst._last_shrink = True
                    Clock.schedule_once(lambda dt: settle(inst, avail), 0.05)
                    return
                inst.text_size = (avail, None)  # 达到下限仍超宽 → 单行省略
                return
            if inst.font_size < max_font and not inst._last_shrink:
                inst.font_size = max_font
                Clock.schedule_once(lambda dt: settle(inst, avail), 0.05)
                return
            inst._last_shrink = False
            # 收敛：text_size 绑定实际宽度（而非 None）——
            # 否则 texture 在 Label 内水平居中，halign="left" 不生效（真机/桌面都居中）
            inst.text_size = (inst.width, None)  # 正常字号单行
        finally:
            inst._fit_lock = False

    def schedule(inst, val):
        if inst._fit_lock:
            return
        if inst._fit_evt is not None:
            inst._fit_evt.cancel()
        inst._fit_evt = Clock.schedule_once(lambda dt: decide(inst, inst.width), 0.12)

    label.bind(size=schedule)


# 注：单笔消费排名不再单独维护行组件——统一复用 RecordRow（传入 rank 参数）。


class SwitchTrack(Widget):
    """轨道 + 白色滑块圆，二者由同一个 widget 在同一个 canvas 上绘制。

    坐标空间说明（这是本组件正确性的核心）：
    Kivy 画布指令的坐标空间 = widget 的父坐标空间；对普通布局树而言，
    widget 的 pos/size 数值就是窗口坐标（布局系统写入含父偏移的累积值）。
    因此本组件只用“自己的 pos/size”推导轨道与滑块的所有画布坐标，
    绝不使用某个子控件的局部坐标去绘制另一个 widget 的画布指令——
    旧实现中滑块画布指令用“相对滑道的局部坐标”定位，实际渲染到窗口原点附近。
    """

    progress = NumericProperty(0.0)  # 0.0=左（是/收入），1.0=右（否/支出）

    def __init__(self, active=False, **kwargs):
        super().__init__(size_hint=(None, 1), width=dp(62), **kwargs)
        self.progress = 0.0 if active else 1.0
        with self.canvas.before:
            self.track_color = Color(*COLOR_BORDER)
            self.track_rect = RoundedRectangle(
                pos=self.pos, size=self.size, radius=[dp(14)] * 4
            )
            Color(1, 1, 1, 1)
            self.knob_rect = RoundedRectangle(
                pos=self.pos, size=self.size, radius=[dp(12)] * 4
            )
        self.bind(pos=self._update_geometry, size=self._update_geometry)
        self.bind(progress=self._update_geometry)
        self._update_geometry(self, None)

    def _update_geometry(self, instance, value):
        """从自身 bounds 推导轨道与滑块（同一画布、同一坐标空间，窗口坐标语义）。"""
        self.track_rect.pos = self.pos
        self.track_rect.size = self.size
        pad = dp(2)
        dia = min(dp(24), self.height - 2 * pad)
        left = self.x + pad
        right = self.right - pad - dia
        kx = left + self.progress * (right - left)
        ky = self.y + (self.height - dia) / 2.0
        self.knob_rect.pos = (kx, ky)
        self.knob_rect.size = (dia, dia)


class IncomeTypeSwitch(ButtonBehavior, BoxLayout):
    """自定义横向椭圆开关：左=是=收入（绿色滑道），右=否=支出（灰色滑道）。

    整个组件（含“是/否”文字与滑道）都可点击切换；
    轨道与滑块由 SwitchTrack 在同一 canvas 上绘制，滑块位置由轨道 bounds
    与归一化状态 progress 推导（0=左端点，1=右端点），约 150ms 平滑切换。
    """
    def __init__(self, active=False, on_toggle=None, **kwargs):
        super().__init__(
            orientation="horizontal", size_hint=(None, None),
            size=(dp(116), dp(34)), spacing=dp(6), padding=[dp(2), dp(2)], **kwargs
        )
        self.active = bool(active)
        self.on_toggle = on_toggle

        self.yes_label = Label(
            text="是", font_size=sp(14), bold=True, size_hint=(None, 1), width=dp(20),
            halign="center", valign="middle"
        )
        self.no_label = Label(
            text="否", font_size=sp(14), bold=True, size_hint=(None, 1), width=dp(20),
            halign="center", valign="middle"
        )
        self.add_widget(self.yes_label)

        # 轨道 + 滑块（同一控件、同一画布、同一坐标空间）
        self.track = SwitchTrack(active=self.active)
        self.add_widget(self.track)
        self.add_widget(self.no_label)

        self.bind(on_release=self._toggle)
        self._refresh_visual()

    def _toggle(self, *args):
        self.active = not self.active
        self._refresh_visual(animate=True)
        if self.on_toggle:
            self.on_toggle(self, self.active)

    def _refresh_visual(self, animate=False):
        # 滑道颜色：收入=柔和绿色，支出=灰色
        self.track.track_color.rgba = COLOR_SUCCESS_LIGHT if self.active else COLOR_BORDER
        # “是/否”激活态文字
        if self.active:
            self.yes_label.color = COLOR_SUCCESS
            self.no_label.color = COLOR_TEXT_SECONDARY
        else:
            self.yes_label.color = COLOR_TEXT_SECONDARY
            self.no_label.color = COLOR_TEXT
        # 滑块状态：0=左（是/收入），1=右（否/支出）；
        # 几何始终由 SwitchTrack 从自身 bounds 推导，动画只改变归一化状态，
        # 轨道移动/缩放时滑块自动保持正确位置。
        target = 0.0 if self.active else 1.0
        if animate and abs(self.track.progress - target) > 0.001:
            Animation(progress=target, duration=0.15).start(self.track)
        else:
            self.track.progress = target


class CategoryPieChart(Widget):
    """使用 Kivy Canvas 绘制按分类消费占比的正圆扇形图，支持点击扇区回调分类。

    只负责绘图与命中判断，点击后的业务 Popup 由 MainScreen 处理。
    """
    def __init__(self, category_stats=None, colors=None, on_category_press=None, **kwargs):
        super().__init__(**kwargs)
        self.category_stats = category_stats or []
        self.colors = colors or CATEGORY_CHART_COLORS
        self.on_category_press = on_category_press
        self.sectors = []
        self.chart_center = (0.0, 0.0)
        self.chart_radius = 0.0
        self.bind(pos=self._redraw, size=self._redraw)
        self._redraw()

    def set_data(self, category_stats):
        self.category_stats = category_stats or []
        self._redraw()

    def _redraw(self, *args):
        self.canvas.clear()
        # 每次重绘重建扇区命中信息，避免尺寸改变后继续使用旧几何数据
        self.sectors = []
        self.chart_center = (0.0, 0.0)
        self.chart_radius = 0.0
        valid_stats = [(category, amount) for category, amount in self.category_stats if amount > 0]
        total = sum(amount for category, amount in valid_stats)
        if total <= 0:
            return

        diameter = min(self.width, self.height)
        if diameter <= 0:
            return

        padding = dp(6)
        diameter = max(0, diameter - padding * 2)
        x = self.x + (self.width - diameter) / 2
        y = self.y + (self.height - diameter) / 2
        radius = diameter / 2.0
        self.chart_center = (x + radius, y + radius)
        self.chart_radius = radius

        start_angle = 0
        with self.canvas:
            for index, (category, amount) in enumerate(valid_stats):
                end_angle = 360 if index == len(valid_stats) - 1 else start_angle + amount / total * 360
                Color(*self.colors[index % len(self.colors)])
                Ellipse(pos=(x, y), size=(diameter, diameter), angle_start=start_angle, angle_end=end_angle)
                self.sectors.append({
                    "category": category,
                    "amount": amount,
                    "start_angle": start_angle,
                    "end_angle": end_angle,
                })
                start_angle = end_angle

    def _get_touch_angle(self, touch_x, touch_y):
        """触点相对实际圆心/半径的命中角度；圆外返回 None。

        角度采用 Kivy Ellipse 的实际语义（0° 在 12 点钟方向、顺时针，
        绘制时 x = r*sin(θ)、y = r*cos(θ)），因此命中角 θ = atan2(dx, dy)。
        """
        if not self.sectors or self.chart_radius <= 0:
            return None
        center_x, center_y = self.chart_center
        dx = touch_x - center_x
        dy = touch_y - center_y
        if dx * dx + dy * dy > self.chart_radius * self.chart_radius:
            return None
        angle = math.degrees(math.atan2(dx, dy))
        if angle < 0:
            angle += 360
        return angle

    def _get_sector_at_angle(self, angle):
        """按半开区间 [start, end) 命中扇区；最后一个扇区包含 360°，一个触点最多命中一个分类。"""
        if angle is None:
            return None
        last_index = len(self.sectors) - 1
        for index, sector in enumerate(self.sectors):
            if index == last_index:
                if sector["start_angle"] <= angle <= sector["end_angle"]:
                    return sector
            elif sector["start_angle"] <= angle < sector["end_angle"]:
                return sector
        return None

    def on_touch_down(self, touch):
        """命中扇区并存在回调时才消费触摸；否则交还父级，避免影响 ScrollView 滚动。"""
        sector = self._get_sector_at_angle(self._get_touch_angle(touch.x, touch.y))
        if sector is not None and self.on_category_press is not None:
            self.on_category_press(sector["category"])
            return True
        return super().on_touch_down(touch)


class MainScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.name = "main"

        self.categories = list(DEFAULT_CATEGORIES)
        self.records = []
        # 页面按需刷新标记：数据变化时才重建列表/统计，普通切换直接显示缓存
        self.records_page_dirty = True
        self.income_page_dirty = True
        self.stats_page_dirty = True
        # 收入功能（默认关闭；income_categories 保存“收入分类”名称集合）
        self.income_enabled = False
        self.income_categories = set()
        self._android_export_bound = False
        self._pending_android_export = None
        self._pending_android_import = False
        self.selected_record_date = datetime.now().date()
        self._success_feedback_event = None

        self.storage_dir = self.get_storage_dir()
        os.makedirs(self.storage_dir, exist_ok=True)

        self.records_path = os.path.join(self.storage_dir, "records.json")
        self.categories_path = os.path.join(self.storage_dir, "categories.json")
        self.settings_path = os.path.join(self.storage_dir, "settings.json")

        self.build_ui()
        self.load_data()
        Clock.schedule_once(lambda dt: self.update_monthly_expense(), 0.1)

    # =========================
    # 基础路径
    # =========================
    def get_storage_dir(self):
        app = App.get_running_app()
        if platform == "android" and app is not None:
            return app.user_data_dir
        return APP_DIR

    def get_export_dir(self):
        # 安卓优先导出到 Download；失败则回退到 app 私有目录
        if platform == "android":
            try:
                from android.storage import primary_external_storage_path
                export_dir = os.path.join(primary_external_storage_path(), "Download")
                os.makedirs(export_dir, exist_ok=True)
                return export_dir
            except Exception:
                pass

        export_dir = os.path.join(self.storage_dir, "exports")
        os.makedirs(export_dir, exist_ok=True)
        return export_dir

    def get_default_import_dir(self):
        export_dir = self.get_export_dir()
        if os.path.exists(export_dir):
            return export_dir
        return self.storage_dir

    # =========================
    # UI 辅助方法
    # =========================
    def _add_rounded_background(self, widget, color, radius, border_color=None):
        """为控件添加随位置和尺寸更新的圆角背景，且不拦截触摸事件。"""
        with widget.canvas.before:
            if border_color is not None:
                Color(*border_color)
                widget.theme_border = RoundedRectangle(pos=widget.pos, size=widget.size, radius=[radius] * 4)
            widget.theme_bg_color = Color(*color)
            inset = dp(1) if border_color is not None else 0
            widget.theme_bg = RoundedRectangle(
                pos=(widget.x + inset, widget.y + inset),
                size=(max(0, widget.width - inset * 2), max(0, widget.height - inset * 2)),
                radius=[max(0, radius - inset)] * 4
            )

        def update_background(instance, value):
            if border_color is not None:
                instance.theme_border.pos = instance.pos
                instance.theme_border.size = instance.size
            instance.theme_bg.pos = (instance.x + inset, instance.y + inset)
            instance.theme_bg.size = (
                max(0, instance.width - inset * 2),
                max(0, instance.height - inset * 2)
            )

        widget.bind(pos=update_background, size=update_background)
        return widget

    def _make_card(self, bg_color=COLOR_CARD_BG, radius=CARD_RADIUS):
        """创建一个带圆角白色背景的卡片容器"""
        card = BoxLayout(size_hint_y=None)
        card.bind(minimum_height=card.setter("height"))

        return self._add_rounded_background(card, bg_color, radius)

    def _make_title_label(self, text, font_size=sp(24), height=dp(48)):
        label = Label(
            text=text,
            font_size=font_size,
            size_hint_y=None,
            height=height,
            color=COLOR_TEXT
        )
        return label

    def _make_section_label(self, text, font_size=sp(18), height=dp(32)):
        label = Label(
            text=text,
            font_size=font_size,
            size_hint_y=None,
            height=height,
            halign="left",
            valign="middle",
            color=COLOR_TEXT
        )
        label.bind(size=lambda inst, val: setattr(inst, "text_size", (val[0], None)))
        return label

    def _make_button(self, text, bg_color, text_color, height=CONTROL_HEIGHT, font_size=sp(17)):
        btn = Button(
            text=text,
            size_hint_y=None,
            height=height,
            font_size=font_size,
            background_normal="",
            background_color=COLOR_TRANSPARENT,
            color=text_color
        )
        return self._add_rounded_background(btn, bg_color, BUTTON_RADIUS)

    def _make_primary_button(self, text, height=PRIMARY_BUTTON_HEIGHT, font_size=sp(18)):
        return self._make_button(text, COLOR_PRIMARY, COLOR_WHITE, height, font_size)

    def _make_secondary_button(self, text, height=CONTROL_HEIGHT, font_size=sp(17)):
        return self._make_button(text, COLOR_PRIMARY_LIGHT, COLOR_PRIMARY, height, font_size)

    def _make_text_button(self, text, height=CONTROL_HEIGHT, font_size=sp(17)):
        return self._make_button(text, COLOR_CARD_BG, COLOR_TEXT_SECONDARY, height, font_size)

    def _make_danger_button(self, text, height=CONTROL_HEIGHT, font_size=sp(17)):
        return self._make_button(text, COLOR_DANGER_LIGHT, COLOR_DANGER, height, font_size)

    def _make_action_button(self, text, color, height=dp(48), font_size=sp(18)):
        if color == COLOR_DANGER:
            return self._make_danger_button(text, height, font_size)
        return self._make_secondary_button(text, height, font_size)

    def _make_text_input(self, hint_text="", input_filter=None, height=CONTROL_HEIGHT, font_size=sp(17), **kwargs):
        ti = TextInput(
            hint_text=hint_text,
            multiline=False,
            input_filter=input_filter,
            size_hint_y=None,
            height=height,
            font_size=font_size,
            background_normal="",
            background_active="",
            background_color=COLOR_WHITE,
            foreground_color=COLOR_TEXT,
            cursor_color=COLOR_PRIMARY,
            selection_color=(0.145, 0.388, 0.922, 0.35),
            hint_text_color=COLOR_TEXT_SECONDARY,
            padding=[dp(12), dp(12), dp(12), dp(10)],
            disabled=False,
            readonly=False,
            write_tab=False,
            **kwargs
        )
        # TextInput 自身绘制不透明白色背景；自定义 Canvas 只绘制边框，
        # 避免部分 Android 设备上填充图形覆盖输入文字的纹理。
        with ti.canvas.after:
            Color(*COLOR_BORDER)
            ti.input_border = Line(
                rounded_rectangle=(ti.x, ti.y, ti.width, ti.height, INPUT_RADIUS),
                width=dp(1)
            )

        def update_input_border(instance, value):
            instance.input_border.rounded_rectangle = (
                instance.x,
                instance.y,
                instance.width,
                instance.height,
                INPUT_RADIUS
            )

        ti.bind(pos=update_input_border, size=update_input_border)
        return ti

    def _make_spinner(self, text, values, height=CONTROL_HEIGHT, font_size=sp(17)):
        spinner = Spinner(
            text=text,
            values=values,
            size_hint_y=None,
            height=height,
            font_size=font_size,
            option_cls=ThemedSpinnerOption,
            background_normal="",
            background_color=COLOR_TRANSPARENT,
            color=COLOR_TEXT
        )
        return self._add_rounded_background(spinner, COLOR_CARD_BG, INPUT_RADIUS, COLOR_BORDER)

    def _make_popup(self, title, content, size_hint, auto_dismiss=False):
        # 清空 Popup 的默认纹理背景，并为内容层提供明确的不透明底色。
        if not hasattr(content, "theme_bg"):
            self._add_rounded_background(content, COLOR_CARD_BG, INPUT_RADIUS)
        return Popup(
            title=title,
            content=content,
            size_hint=size_hint,
            auto_dismiss=auto_dismiss,
            background="",
            background_color=COLOR_CARD_BG,
            separator_color=COLOR_BORDER,
            title_color=COLOR_TEXT,
            title_size=sp(20)
        )

    # =========================
    # UI
    # =========================
    def build_ui(self):
        """构建共享同一业务状态的四页移动端导航骨架。"""
        root = BoxLayout(orientation="vertical")
        with root.canvas.before:
            Color(*COLOR_PAGE_BG)
            root.bg = RoundedRectangle(radius=[0] * 4, pos=root.pos, size=root.size)

        def update_root_bg(instance, value):
            instance.bg.pos = instance.pos
            instance.bg.size = instance.size

        root.bind(pos=update_root_bg, size=update_root_bg)
        self.page_manager = ScreenManager()
        self._nav_buttons = {}
        self.page_manager.add_widget(self._build_accounting_screen())
        self.page_manager.add_widget(self._build_records_screen())
        self.page_manager.add_widget(self._build_income_screen())
        self.page_manager.add_widget(self._build_stats_screen())
        self.page_manager.add_widget(self._build_settings_screen())
        root.add_widget(self.page_manager)
        self.add_widget(root)
        Window.bind(on_keyboard=self._handle_back_key)
        self.bind(parent=self._unbind_window_keyboard)
        self.switch_page("accounting")

    def _make_page_header(self, title, action_text=None, action_callback=None):
        header = BoxLayout(size_hint_y=None, height=dp(52), spacing=dp(8))
        title_label = self._make_title_label(title, height=dp(52))
        title_label.halign = "left"
        title_label.valign = "middle"
        title_label.bind(size=lambda inst, val: setattr(inst, "text_size", val))
        header.add_widget(title_label)
        if action_text:
            action = self._make_text_button(action_text, height=dp(48), font_size=sp(16))
            action.size_hint_x = None
            action.width = dp(80)
            action.bind(on_press=action_callback)
            header.add_widget(action)
        return header

    def _make_page_scroll(self, content):
        scroll = ScrollView(size_hint=(1, 1))
        content.size_hint_y = None
        content.bind(minimum_height=content.setter("height"))
        scroll.add_widget(content)
        return scroll

    def _make_bottom_navigation(self, selected_page):
        nav = BoxLayout(
            size_hint_y=None, height=dp(64), spacing=dp(8),
            padding=[PAGE_PADDING, dp(8), PAGE_PADDING, dp(8)]
        )
        self._add_rounded_background(nav, COLOR_CARD_BG, 0, COLOR_BORDER)
        nav_pages = [("accounting", "记账"), ("records", "支出")]
        if self.income_enabled:
            nav_pages.append(("income", "收入"))
        nav_pages.append(("stats", "统计"))
        for page_name, text in nav_pages:
            selected = page_name == selected_page
            button = self._make_button(
                text,
                COLOR_PRIMARY if selected else COLOR_PRIMARY_LIGHT,
                COLOR_WHITE if selected else COLOR_TEXT_SECONDARY,
                height=dp(48),
                font_size=sp(16)
            )
            button.bind(on_press=lambda instance, target=page_name: self.switch_page(target))
            nav.add_widget(button)
            self._nav_buttons[(selected_page, page_name)] = button
        return nav

    def _refresh_all_bottom_navigations(self):
        """收入开关变化后重建四个主页面底栏（记账/支出/收入/统计）。"""
        for page_name in ("accounting", "records", "income", "stats"):
            screen = self.page_manager.get_screen(page_name)
            page = screen.children[0]
            if page.children:
                page.remove_widget(page.children[0])  # 旧底栏（始终是最后添加的第一个子项）
            page.add_widget(self._make_bottom_navigation(page_name))

    def _build_accounting_screen(self):
        screen = Screen(name="accounting")
        page = BoxLayout(orientation="vertical")
        content = BoxLayout(
            orientation="vertical", spacing=CARD_SPACING,
            padding=[PAGE_PADDING, PAGE_PADDING, PAGE_PADDING, dp(24)]
        )
        content.add_widget(self._make_page_header(
            "个人记账", "设置", lambda instance: self.switch_page("settings")
        ))

        expense_card = self._make_card()
        expense_layout = BoxLayout(
            orientation="vertical", padding=[CARD_PADDING, dp(16), CARD_PADDING, dp(14)],
            spacing=dp(4), size_hint_y=None, height=dp(132)
        )
        expense_layout.add_widget(self._make_section_label("本月总支出", font_size=sp(16), height=dp(24)))
        self.monthly_expense_label = Label(
            text="0.00 元", font_size=sp(34), size_hint_y=None, height=dp(58), color=COLOR_PRIMARY
        )
        expense_layout.add_widget(self.monthly_expense_label)
        self.record_count_label = Label(
            text="记录数：0", font_size=sp(16), size_hint_y=None,
            height=dp(24), color=COLOR_TEXT_SECONDARY
        )
        expense_layout.add_widget(self.record_count_label)
        expense_card.add_widget(expense_layout)
        content.add_widget(expense_card)

        form_card = self._make_card()
        form_layout = BoxLayout(
            orientation="vertical", spacing=dp(10),
            padding=[CARD_PADDING] * 4, size_hint_y=None
        )
        form_layout.bind(minimum_height=form_layout.setter("height"))
        form_layout.add_widget(self._make_section_label("金额（元）："))
        self.amount_input = self._make_text_input(hint_text="0.00", input_filter="float")
        form_layout.add_widget(self.amount_input)
        form_layout.add_widget(self._make_section_label("分类："))
        self.category_spinner = self._make_spinner(text="饮食正餐", values=self.categories)
        form_layout.add_widget(self.category_spinner)
        form_layout.add_widget(self._make_section_label("消费备注："))
        self.name_input = self._make_text_input(hint_text="例如：午餐")
        form_layout.add_widget(self.name_input)

        date_layout = BoxLayout(size_hint_y=None, height=dp(64), spacing=dp(8))
        date_text_box = BoxLayout(orientation="vertical", spacing=dp(2))
        date_text_box.add_widget(self._make_section_label("日期", height=dp(24)))
        self.record_date_label = Label(
            color=COLOR_TEXT, font_size=sp(16), halign="left", valign="middle"
        )
        self.record_date_label.bind(
            size=lambda inst, val: setattr(inst, "text_size", val)
        )
        date_text_box.add_widget(self.record_date_label)
        date_layout.add_widget(date_text_box)
        change_date_btn = self._make_text_button("修改", height=dp(48), font_size=sp(16))
        change_date_btn.size_hint_x = None
        change_date_btn.width = dp(72)
        change_date_btn.bind(on_press=self.open_record_date_popup)
        date_layout.add_widget(change_date_btn)
        self._update_record_date_label()
        form_layout.add_widget(date_layout)
        record_btn = self._make_primary_button("记一笔", height=PRIMARY_BUTTON_HEIGHT, font_size=sp(19))
        record_btn.bind(on_press=self.record_bill)
        form_layout.add_widget(record_btn)
        self.record_success_label = Label(
            text="", color=COLOR_SUCCESS, font_size=sp(15),
            size_hint_y=None, height=dp(28), opacity=0,
            halign="center", valign="middle"
        )
        self.record_success_label.bind(
            size=lambda inst, val: setattr(inst, "text_size", val)
        )
        form_layout.add_widget(self.record_success_label)
        form_card.add_widget(form_layout)
        content.add_widget(form_card)

        page.add_widget(self._make_page_scroll(content))
        page.add_widget(self._make_bottom_navigation("accounting"))
        screen.add_widget(page)
        return screen

    def _build_records_screen(self):
        screen = Screen(name="records")
        page = BoxLayout(orientation="vertical")
        content = BoxLayout(
            orientation="vertical", spacing=CARD_SPACING,
            padding=[PAGE_PADDING, PAGE_PADDING, PAGE_PADDING, dp(24)]
        )
        content.add_widget(self._make_page_header(
            "支出", "设置", lambda instance: self.switch_page("settings")
        ))
        records_card = self._make_card()
        self.records_list = GridLayout(
            cols=1, spacing=dp(8), padding=[CARD_PADDING] * 4, size_hint_y=None
        )
        self.records_list.bind(minimum_height=self.records_list.setter("height"))
        records_card.add_widget(self.records_list)
        content.add_widget(records_card)
        page.add_widget(self._make_page_scroll(content))
        page.add_widget(self._make_bottom_navigation("records"))
        screen.add_widget(page)
        return screen

    def _build_income_screen(self):
        screen = Screen(name="income")
        page = BoxLayout(orientation="vertical")
        content = BoxLayout(
            orientation="vertical", spacing=CARD_SPACING,
            padding=[PAGE_PADDING, PAGE_PADDING, PAGE_PADDING, dp(24)]
        )
        content.add_widget(self._make_page_header(
            "收入", "设置", lambda instance: self.switch_page("settings")
        ))
        income_card = self._make_card()
        self.income_list = GridLayout(
            cols=1, spacing=dp(8), padding=[CARD_PADDING] * 4, size_hint_y=None
        )
        self.income_list.bind(minimum_height=self.income_list.setter("height"))
        income_card.add_widget(self.income_list)
        content.add_widget(income_card)
        page.add_widget(self._make_page_scroll(content))
        page.add_widget(self._make_bottom_navigation("income"))
        screen.add_widget(page)
        return screen

    def _build_stats_screen(self):
        screen = Screen(name="stats")
        page = BoxLayout(orientation="vertical")
        content = BoxLayout(
            orientation="vertical", spacing=CARD_SPACING,
            padding=[PAGE_PADDING, PAGE_PADDING, PAGE_PADDING, dp(24)]
        )
        content.add_widget(self._make_page_header(
            "统计", "设置", lambda instance: self.switch_page("settings")
        ))

        controls_card = self._make_card()
        controls = BoxLayout(
            orientation="vertical", spacing=dp(10), padding=[CARD_PADDING] * 4, size_hint_y=None
        )
        controls.bind(minimum_height=controls.setter("height"))
        controls.add_widget(self._make_section_label("统计类型", font_size=sp(16), height=dp(26)))
        mode_selector = BoxLayout(size_hint_y=None, height=CONTROL_HEIGHT, spacing=dp(8))
        self.stats_month_button = self._make_primary_button("月度", height=CONTROL_HEIGHT, font_size=sp(16))
        self.stats_year_button = self._make_text_button("年度", height=CONTROL_HEIGHT, font_size=sp(16))
        self.stats_month_button.bind(on_press=lambda instance: self._set_stats_mode("month"))
        self.stats_year_button.bind(on_press=lambda instance: self._set_stats_mode("year"))
        mode_selector.add_widget(self.stats_month_button)
        mode_selector.add_widget(self.stats_year_button)
        controls.add_widget(mode_selector)
        controls.add_widget(self._make_section_label("统计周期", font_size=sp(16), height=dp(26)))
        self.stats_period_spinner = self._make_spinner("", ())
        self.stats_period_spinner.bind(text=self._on_stats_period_changed)
        controls.add_widget(self.stats_period_spinner)
        controls_card.add_widget(controls)
        content.add_widget(controls_card)

        self.stats_mode = "month"
        self._updating_stats_period = False
        self.stats_results = BoxLayout(orientation="vertical", spacing=CARD_SPACING, size_hint_y=None)
        self.stats_results.bind(minimum_height=self.stats_results.setter("height"))
        content.add_widget(self.stats_results)
        page.add_widget(self._make_page_scroll(content))
        page.add_widget(self._make_bottom_navigation("stats"))
        screen.add_widget(page)
        return screen

    def _build_settings_screen(self):
        screen = Screen(name="settings")
        content = BoxLayout(
            orientation="vertical", spacing=CARD_SPACING,
            padding=[PAGE_PADDING, PAGE_PADDING, PAGE_PADDING, dp(24)]
        )
        content.add_widget(self._make_page_header(
            "设置", "返回", lambda instance: self.switch_page("accounting")
        ))

        overview_card = self._make_card()
        overview_box = BoxLayout(
            orientation="vertical", spacing=dp(8), padding=[CARD_PADDING] * 4, size_hint_y=None
        )
        overview_box.bind(minimum_height=overview_box.setter("height"))
        self.settings_records_count_label = self._make_section_label("账单记录：0 条", font_size=sp(17), height=dp(28))
        self.settings_categories_count_label = self._make_section_label("分类数量：0 个", font_size=sp(17), height=dp(28))
        settings_storage_tip = Label(
            text="数据保存在本机应用目录，卸载应用前请先导出完整备份。",
            color=COLOR_TEXT_SECONDARY, font_size=sp(14), halign="left", valign="middle",
            size_hint_y=None, height=dp(44)
        )
        settings_storage_tip.bind(size=lambda inst, val: setattr(inst, "text_size", (val[0], None)))
        overview_box.add_widget(self.settings_records_count_label)
        overview_box.add_widget(self.settings_categories_count_label)
        overview_box.add_widget(settings_storage_tip)
        overview_card.add_widget(overview_box)
        content.add_widget(overview_card)

        income_card = self._make_card()
        income_box = BoxLayout(
            orientation="vertical", spacing=dp(10), padding=[CARD_PADDING] * 4, size_hint_y=None
        )
        income_box.bind(minimum_height=income_box.setter("height"))
        income_row = BoxLayout(size_hint_y=None, height=CONTROL_HEIGHT, spacing=dp(8))
        income_row.add_widget(self._make_section_label("启用收入功能", font_size=sp(17), height=CONTROL_HEIGHT))
        self.income_switch = IncomeTypeSwitch(
            active=self.income_enabled, on_toggle=self._on_income_switch_changed,
            pos_hint={"center_y": 0.5}
        )
        income_row.add_widget(self.income_switch)
        income_box.add_widget(income_row)
        income_box.add_widget(Label(
            text="开启后底部导航显示“收入”，记账页可选择收入分类",
            color=COLOR_TEXT_SECONDARY, font_size=sp(13), halign="left", valign="middle",
            size_hint_y=None, height=dp(34)
        ))
        income_card.add_widget(income_box)
        content.add_widget(income_card)

        sections = (
            ("分类管理", "管理记账时可选择的消费分类", (("管理分类", self.show_categories, False),)),
            ("数据管理", "导入数据：从 JSON、CSV 或 Excel 恢复账单\n导出数据：生成账单表格或完整备份", (
                ("导入数据", self.import_data_popup, False),
                ("导出数据", self.export_data, False),
            )),
            ("危险操作", "清空所有账单记录，不删除分类", (
                ("清空所有记录", self.clear_all_records, True),
            )),
        )
        for title, description, actions in sections:
            content.add_widget(self._make_section_label(title))
            card = self._make_card()
            box = BoxLayout(
                orientation="vertical", spacing=dp(10), padding=[CARD_PADDING] * 4, size_hint_y=None
            )
            box.bind(minimum_height=box.setter("height"))
            desc_label = Label(
                text=description, color=COLOR_TEXT_SECONDARY, font_size=sp(14),
                halign="left", valign="middle", size_hint_y=None, height=dp(44)
            )
            desc_label.bind(size=lambda inst, val: setattr(inst, "text_size", (val[0], None)))
            box.add_widget(desc_label)
            for text, callback, danger in actions:
                button = (self._make_danger_button if danger else self._make_secondary_button)(text)
                button.bind(on_press=callback)
                box.add_widget(button)
            card.add_widget(box)
            content.add_widget(card)
        screen.add_widget(self._make_page_scroll(content))
        self.refresh_settings_page()
        return screen

    def refresh_settings_page(self):
        if hasattr(self, "settings_records_count_label"):
            self.settings_records_count_label.text = f"账单记录：{len(self.records)} 条"
        if hasattr(self, "settings_categories_count_label"):
            self.settings_categories_count_label.text = f"分类数量：{len(self.categories)} 个"

    def _on_income_switch_changed(self, instance, value):
        """收入开关：立即生效（底栏/分类/统计联动），数据不丢失。"""
        if self.income_enabled == value:
            return
        self.income_enabled = value
        self._save_settings()
        self._refresh_all_bottom_navigations()
        self._sync_category_spinner()
        self.income_page_dirty = True
        self.stats_page_dirty = True
        if not value and self.page_manager.current == "income":
            self.switch_page("records")

    def switch_page(self, page_name):
        """切换页面；按页面顺序自动选择左右动画方向，并保留原有刷新逻辑。

        性能：记录/收入/统计页采用按需刷新（dirty flag）——数据未变化时
        直接显示已有页面，避免每次切换都重建列表与统计。
        """
        current_page = self.page_manager.current
        if page_name == "records":
            if self.records_page_dirty:
                self.refresh_records_page()
                self.records_page_dirty = False
        elif page_name == "income":
            if self.income_page_dirty:
                self.refresh_income_page()
                self.income_page_dirty = False
        elif page_name == "stats":
            if self.stats_page_dirty:
                self.refresh_stats_page()
                self.stats_page_dirty = False
        elif page_name == "settings":
            self.refresh_settings_page()

        if page_name == current_page:
            return

        direction = self._get_page_transition_direction(current_page, page_name)
        if direction is not None:
            self.page_manager.transition.direction = direction
        self.page_manager.current = page_name

    def _get_page_transition_direction(self, current_page, target_page):
        """根据页面顺序判定切换动画方向，返回 None 表示无需动画。"""
        if current_page == target_page:
            return None
        current_index = PAGE_ORDER.get(current_page)
        target_index = PAGE_ORDER.get(target_page)
        if current_index is not None and target_index is not None:
            return "left" if target_index > current_index else "right"
        if current_page == "accounting" and target_page == "settings":
            return "left"
        if current_page == "settings" and target_page == "accounting":
            return "right"
        return "left"

    def refresh_records_page(self):
        """按既有排序规则刷新支出页最近 50 条记录（只显示记录类型为“支出”）。"""
        self.sort_records()
        self.records_list.clear_widgets()
        display_records = [
            record for record in self.records
            if isinstance(record, dict) and self._get_record_type(record) == "支出"
        ][:50]
        if not display_records:
            empty = BoxLayout(orientation="vertical", spacing=dp(8), size_hint_y=None, height=dp(164))
            empty.add_widget(Label(
                text="暂无记录", color=COLOR_TEXT, size_hint_y=None,
                height=dp(40), font_size=sp(20), bold=True
            ))
            empty.add_widget(Label(
                text="记一笔消费后会显示在这里", color=COLOR_TEXT_SECONDARY,
                size_hint_y=None, height=dp(36), font_size=sp(15)
            ))
            go_accounting = self._make_secondary_button("去记账")
            go_accounting.bind(on_press=lambda instance: self.switch_page("accounting"))
            empty.add_widget(go_accounting)
            self.records_list.add_widget(empty)
            return
        for record in display_records:
            self.records_list.add_widget(RecordRow(record, self.show_record_detail))

    def refresh_income_page(self):
        """刷新收入页：只显示记录类型为“收入”的记录（最多 50 条）。"""
        self.sort_records()
        self.income_list.clear_widgets()
        display_records = [
            record for record in self.records
            if isinstance(record, dict) and self._get_record_type(record) == "收入"
        ][:50]
        if not display_records:
            empty = BoxLayout(orientation="vertical", spacing=dp(8), size_hint_y=None, height=dp(164))
            empty.add_widget(Label(
                text="暂无收入记录", color=COLOR_TEXT, size_hint_y=None,
                height=dp(40), font_size=sp(20), bold=True
            ))
            empty.add_widget(Label(
                text="开启收入功能后，使用收入分类记账即可", color=COLOR_TEXT_SECONDARY,
                size_hint_y=None, height=dp(36), font_size=sp(15)
            ))
            self.income_list.add_widget(empty)
            return
        for record in display_records:
            self.income_list.add_widget(RecordRow(record, self.show_record_detail))

    def show_record_detail(self, record):
        """显示完整记录，并从当前记录对象发起安全的单条删除流程。"""
        if not isinstance(record, dict):
            self.show_popup("提示", "这条记录格式异常，无法查看详情。")
            return
        content = BoxLayout(orientation="vertical", spacing=dp(12), padding=CARD_PADDING)
        scroll = ScrollView(size_hint=(1, 1), do_scroll_x=False)
        details = BoxLayout(orientation="vertical", spacing=dp(8), size_hint_y=None)
        details.bind(minimum_height=details.setter("height"))
        fields = [
            ("消费备注", str(record.get("姓名/备注", ""))),
            ("分类", str(record.get("分类", ""))),
            ("金额", self._format_record_amount(record)),
            ("日期", str(record.get("日期", ""))),
        ]
        record_time = str(record.get("记录时间", "")).strip()
        if record_time:
            fields.append(("记录时间", record_time))

        for field_name, value in fields:
            label = Label(
                text=f"[color=6B7280]{field_name}[/color]\n{escape_markup(value)}", markup=True,
                color=COLOR_TEXT, font_size=sp(16), halign="left", valign="middle",
                size_hint_y=None, height=dp(58)
            )
            label.bind(width=lambda inst, val: setattr(inst, "text_size", (val, None)))
            label.bind(texture_size=lambda inst, val: setattr(inst, "height", max(dp(58), val[1] + dp(8))))
            details.add_widget(label)
        scroll.add_widget(details)
        content.add_widget(scroll)

        buttons = BoxLayout(size_hint_y=None, height=CONTROL_HEIGHT, spacing=dp(10))
        close_button = self._make_text_button("关闭")
        delete_button = self._make_danger_button("删除记录")
        buttons.add_widget(close_button)
        buttons.add_widget(delete_button)
        content.add_widget(buttons)

        detail_popup = self._make_popup("记录详情", content, (0.9, 0.72))
        close_button.bind(on_press=detail_popup.dismiss)
        delete_button.bind(
            on_press=lambda instance: self.confirm_record_deletion(record, detail_popup)
        )
        detail_popup.open()

    def _format_record_amount(self, record):
        try:
            return f"{float(record.get('金额', 0)):.2f} 元"
        except (TypeError, ValueError):
            return f"{record.get('金额', '')} 元"

    def confirm_record_deletion(self, record, detail_popup):
        """二次确认后按对象身份删除，避免连续操作使用过期索引。"""
        def do_delete(instance):
            for index, current_record in enumerate(self.records):
                if current_record is record:
                    del self.records[index]
                    self.save_data()
                    self.update_monthly_expense()
                    detail_popup.dismiss()
                    if self._get_record_type(record) == "收入":
                        self.income_page_dirty = True
                    else:
                        self.records_page_dirty = True
                    self.stats_page_dirty = True
                    self.refresh_records_page()
                    return

            detail_popup.dismiss()
            self.show_popup("提示", "这条记录已不存在。")
            self.refresh_records_page()

        self.show_confirm_popup(
            "确认删除", "确定要删除这条记录吗？此操作不可撤销。", do_delete
        )

    def _handle_back_key(self, window, key, *args):
        if key == 27 and getattr(self, "page_manager", None) is not None:
            if self.page_manager.current != "accounting":
                self.switch_page("accounting")
                return True
        return False

    def _unbind_window_keyboard(self, instance, parent):
        if parent is None:
            Window.unbind(on_keyboard=self._handle_back_key)
            if self._success_feedback_event is not None:
                self._success_feedback_event.cancel()
                self._success_feedback_event = None

    def make_card(self):
        """保留原有方法，供其他可能调用的地方使用"""
        return self._make_card()

    def make_field_label(self, text):
        """保留原有方法，供其他可能调用的地方使用"""
        return self._make_section_label(text)

    # =========================
    # 数据处理
    # =========================
    def _has_exportable_records(self):
        return any(isinstance(record, dict) for record in self.records)

    def _get_record_sort_datetime(self, record):
        """统一记录时间排序键：优先“记录时间”，其次“日期”，解析失败回退 datetime.min。"""
        if not isinstance(record, dict):
            return datetime.min
        record_time = str(record.get("记录时间", "")).strip()
        date_str = str(record.get("日期", "")).strip()

        try:
            if record_time:
                return datetime.strptime(record_time, "%Y-%m-%d %H:%M:%S")
        except Exception:
            pass

        try:
            if date_str:
                return datetime.strptime(date_str, "%Y-%m-%d")
        except Exception:
            pass

        return datetime.min

    def sort_records(self):
        self.records.sort(key=self._get_record_sort_datetime, reverse=True)

    def _ensure_record_id(self, record):
        """确保记录有稳定的唯一标识（“记录ID”）；缺失时生成 UUID 并写入记录。"""
        if not isinstance(record, dict):
            return None
        rid = record.get("记录ID")
        if isinstance(rid, str) and rid.strip():
            return rid.strip()
        rid = uuid.uuid4().hex
        record["记录ID"] = rid
        return rid

    def _get_record_type(self, record):
        """记录自身的类型：收入 / 支出。历史记录缺失时默认为支出。

        类型保存在记录上（而非由分类配置推断），保证分类属性变更不会
        改变历史记录的收入/支出身份。
        """
        record_type = str(record.get("记录类型", "支出")).strip()
        return record_type if record_type in ("收入", "支出") else "支出"

    def _is_income_category(self, category):
        """分类是否为“收入分类”（决定新建记录时的默认类型）。"""
        return category in self.income_categories

    def _load_settings(self):
        """读取 settings.json（收入开关与收入分类集合）；缺失或损坏时安全回退。"""
        self.income_enabled = False
        self.income_categories = set()
        try:
            if os.path.exists(self.settings_path):
                with open(self.settings_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, dict):
                    self.income_enabled = bool(data.get("income_enabled", False))
                    cats = data.get("income_categories", [])
                    if isinstance(cats, list):
                        self.income_categories = {
                            str(c).strip() for c in cats if isinstance(c, str) and c.strip()
                        }
        except Exception:
            self.income_enabled = False
            self.income_categories = set()

    def _save_settings(self):
        """持久化收入开关与收入分类集合。"""
        try:
            data = {
                "income_enabled": self.income_enabled,
                "income_categories": sorted(self.income_categories),
            }
            with open(self.settings_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            self.show_popup("错误", f"保存设置失败：\n{str(e)}")

    def load_data(self):
        try:
            if os.path.exists(self.records_path):
                with open(self.records_path, "r", encoding="utf-8") as f:
                    loaded_records = json.load(f)
                    self.records = loaded_records if isinstance(loaded_records, list) else []

            loaded_categories = self._load_categories_from_file()
            if loaded_categories is not None:
                # 文件中的分类列表才是用户真实配置（含顺序与删除结果），
                # 不要把默认分类重新补进用户已经保存的配置。
                self.categories = loaded_categories

            self._load_settings()

            # 旧数据迁移（一次性并持久化，幂等）：
            # 1) 缺少唯一标识的记录补发稳定记录ID；
            # 2) 缺少“记录类型”的旧记录默认迁移为“支出”（App 原本只有支出）。
            needs_migration = False
            for record in self.records:
                if not isinstance(record, dict):
                    continue
                if not (isinstance(record.get("记录ID"), str) and record.get("记录ID", "").strip()):
                    self._ensure_record_id(record)
                    needs_migration = True
                if self._get_record_type(record) != record.get("记录类型"):
                    record["记录类型"] = self._get_record_type(record)
                    needs_migration = True

            self.sort_records()
            self._sync_category_spinner()

            if needs_migration:
                self.save_data()
        except Exception as e:
            self.records = []
            self.show_popup("提示", f"读取本地数据失败：\n{str(e)}")

    def _load_categories_from_file(self):
        """读取 categories.json 并清洗为有效分类列表。

        清洗规则：只接受字符串、strip 前后空格、丢弃空字符串、
        去除重复值（保留第一次出现的位置）、保持文件中的原始顺序。

        文件缺失、解析失败、类型错误或清洗后没有有效分类时返回 None，
        由调用方回退到默认分类。
        """
        if not os.path.exists(self.categories_path):
            return None
        try:
            with open(self.categories_path, "r", encoding="utf-8") as f:
                loaded = json.load(f)
        except Exception:
            return None
        if not isinstance(loaded, list):
            return None
        cleaned = []
        seen = set()
        for item in loaded:
            if not isinstance(item, str):
                continue
            name = item.strip()
            if not name or name in seen:
                continue
            seen.add(name)
            cleaned.append(name)
        if not cleaned:
            return None
        return cleaned

    def save_data(self):
        try:
            self.sort_records()
            with open(self.records_path, "w", encoding="utf-8") as f:
                json.dump(self.records, f, ensure_ascii=False, indent=2)

            with open(self.categories_path, "w", encoding="utf-8") as f:
                json.dump(self.categories, f, ensure_ascii=False, indent=2)
        except Exception as e:
            self.show_popup("错误", f"保存数据失败：\n{str(e)}")

    def update_monthly_expense(self):
        current_month = datetime.now().strftime("%Y-%m")
        total = 0.0
        count = 0

        for record in self.records:
            if not isinstance(record, dict):
                continue
            if self._get_record_type(record) != "支出":
                continue
            try:
                record_date = datetime.strptime(str(record.get("日期", "")), "%Y-%m-%d")
                if record_date.strftime("%Y-%m") == current_month:
                    total += float(record.get("金额", 0))
                    count += 1
            except Exception:
                continue

        self.monthly_expense_label.text = f"{total:.2f} 元"
        self.record_count_label.text = f"记录数：{count}"

    # =========================
    # 记账
    # =========================
    def _update_record_date_label(self):
        date_text = self.selected_record_date.strftime("%Y-%m-%d")
        prefix = "今天" if self.selected_record_date == datetime.now().date() else "已选"
        self.record_date_label.text = f"{prefix} · {date_text}"

    def open_record_date_popup(self, instance):
        """使用当前选中日期初始化临时控件，取消时不写回页面状态。"""
        current_date = self.selected_record_date
        content = BoxLayout(
            orientation="vertical", spacing=CARD_SPACING, padding=CARD_PADDING
        )
        date_fields = GridLayout(cols=3, spacing=dp(8), size_hint_y=None, height=dp(82))
        year_input = self._make_text_input(
            text=str(current_date.year), input_filter="int"
        )
        month_spinner = self._make_spinner(
            str(current_date.month), [str(i) for i in range(1, 13)]
        )
        day_spinner = self._make_spinner(
            str(current_date.day), [str(i) for i in range(1, 32)]
        )
        for label_text, field in (
            ("年", year_input), ("月", month_spinner), ("日", day_spinner)
        ):
            field_box = BoxLayout(orientation="vertical", spacing=dp(4))
            field_box.add_widget(Label(
                text=label_text, color=COLOR_TEXT_SECONDARY, font_size=sp(15),
                size_hint_y=None, height=dp(24)
            ))
            field_box.add_widget(field)
            date_fields.add_widget(field_box)
        content.add_widget(date_fields)

        actions = BoxLayout(size_hint_y=None, height=dp(48), spacing=dp(8))
        today_button = self._make_secondary_button("今天")
        cancel_button = self._make_text_button("取消")
        confirm_button = self._make_primary_button("确定", height=dp(48), font_size=sp(17))
        actions.add_widget(today_button)
        actions.add_widget(cancel_button)
        actions.add_widget(confirm_button)
        content.add_widget(actions)

        popup = self._make_popup("修改日期", content, (0.92, 0.38))

        def use_today(button):
            today = datetime.now().date()
            year_input.text = str(today.year)
            month_spinner.text = str(today.month)
            day_spinner.text = str(today.day)

        def confirm_date(button):
            try:
                year_text = year_input.text.strip()
                if not year_text:
                    self.show_popup("错误", "请输入年份。")
                    return
                year = int(year_text)
                if year < 1900 or year > 9999:
                    self.show_popup("错误", "请输入合理的年份，例如 2026。")
                    return
                selected = datetime(
                    year, int(month_spinner.text), int(day_spinner.text)
                ).date()
            except (TypeError, ValueError):
                self.show_popup("错误", "日期无效，请检查年月日。")
                return
            self.selected_record_date = selected
            self._update_record_date_label()
            popup.dismiss()

        today_button.bind(on_press=use_today)
        cancel_button.bind(on_press=popup.dismiss)
        confirm_button.bind(on_press=confirm_date)
        popup.open()

    def _hide_record_success(self, dt):
        if not getattr(self, "record_success_label", None):
            self._success_feedback_event = None
            return
        self.record_success_label.text = ""
        self.record_success_label.opacity = 0
        self._success_feedback_event = None

    def _show_record_success(self, amount, note):
        if self._success_feedback_event is not None:
            self._success_feedback_event.cancel()
        self.record_success_label.text = f"已记录 {amount:.2f} 元 · {note}"
        self.record_success_label.opacity = 1
        self._success_feedback_event = Clock.schedule_once(
            self._hide_record_success, 2
        )

    def record_bill(self, instance):
        note = self.name_input.text.strip()
        category = self.category_spinner.text.strip()
        amount_text = self.amount_input.text.strip()

        if not note:
            self.show_popup("错误", "请输入消费备注。")
            return

        if not amount_text:
            self.show_popup("错误", "请输入金额。")
            return

        try:
            amount = round(float(amount_text), 2)
            if amount <= 0:
                raise ValueError
        except Exception:
            self.show_popup("错误", "请输入有效的正数金额。")
            return

        date_str = self.selected_record_date.strftime("%Y-%m-%d")

        record = {
            "记录类型": "收入" if self._is_income_category(category) else "支出",
            "姓名/备注": note,
            "分类": category,
            "金额": amount,
            "日期": date_str,
            "记录时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        self._ensure_record_id(record)

        self.records.append(record)
        self.save_data()
        self.update_monthly_expense()
        if self._get_record_type(record) == "收入":
            self.income_page_dirty = True
        else:
            self.records_page_dirty = True
        self.stats_page_dirty = True

        self.name_input.text = ""
        self.amount_input.text = ""
        self.selected_record_date = datetime.now().date()
        self._update_record_date_label()
        self._show_record_success(amount, note)

    # =========================
    # 统计
    # =========================
    def _set_stats_mode(self, mode):
        """切换页面内统计类型，并立即显示该类型的当前周期。"""
        if mode not in ("month", "year"):
            return
        self.stats_mode = mode
        month_selected = mode == "month"
        self.stats_month_button.theme_bg_color.rgba = (
            COLOR_PRIMARY if month_selected else COLOR_CARD_BG
        )
        self.stats_month_button.color = COLOR_WHITE if month_selected else COLOR_TEXT_SECONDARY
        self.stats_year_button.theme_bg_color.rgba = (
            COLOR_CARD_BG if month_selected else COLOR_PRIMARY
        )
        self.stats_year_button.color = COLOR_TEXT_SECONDARY if month_selected else COLOR_WHITE
        self.refresh_stats_page(reset_period=True)

    def _on_stats_period_changed(self, spinner, period):
        """Spinner 的唯一绑定入口，避免刷新选项时重复触发重建。"""
        if not self._updating_stats_period and period:
            self.refresh_stats_page()

    def _get_stats_periods(self):
        """在既有周期提取结果前补充当前周期，且保持历史周期倒序。"""
        if self.stats_mode == "year":
            current = datetime.now().strftime("%Y")
            available = self.get_available_years()
        else:
            current = datetime.now().strftime("%Y-%m")
            available = self.get_available_months()
        return [current] + [period for period in available if period != current]

    def _build_income_summary_card(self, period, income_total, expense_total):
        """收入功能开启时统计页顶部的“收支概览”卡片（标题在卡片内部）。"""
        card = self._make_card()
        summary = BoxLayout(
            orientation="vertical", spacing=dp(3), padding=[CARD_PADDING] * 4,
            size_hint_y=None, height=dp(140)
        )
        period_suffix = "年度收支统计" if self.stats_mode == "year" else "月度收支统计"
        summary.add_widget(self._make_stats_label(
            f"{period} {period_suffix}", sp(14), dp(26), COLOR_TEXT_SECONDARY, halign="left"
        ))
        summary.add_widget(self._make_stats_label(
            f"收入：{income_total:.2f} 元", sp(17), dp(28), COLOR_SUCCESS, halign="left"
        ))
        summary.add_widget(self._make_stats_label(
            f"支出：{expense_total:.2f} 元", sp(17), dp(28), COLOR_PRIMARY, halign="left"
        ))
        diff = income_total - expense_total
        if diff > 0:
            diff_color = COLOR_SUCCESS
        elif diff < 0:
            diff_color = COLOR_DANGER
        else:
            diff_color = COLOR_TEXT_SECONDARY
        summary.add_widget(self._make_stats_label(
            f"收支差：{diff:.2f} 元", sp(17), dp(28), diff_color, halign="left"
        ))
        card.add_widget(summary)
        return card

    def _build_stats_summary_card(self, period, total, record_count):
        card = self._make_card()
        summary = BoxLayout(
            orientation="vertical", spacing=dp(4), padding=[CARD_PADDING] * 4,
            size_hint_y=None, height=dp(142)
        )
        period_suffix = "年度统计" if self.stats_mode == "year" else "月度统计"
        summary.add_widget(self._make_stats_label(
            f"{period}  {period_suffix}", sp(15), dp(24), COLOR_TEXT_SECONDARY, halign="left"
        ))
        summary.add_widget(self._make_stats_label(
            f"{total:.2f} 元", sp(32), dp(58), COLOR_PRIMARY, halign="left"
        ))
        summary.add_widget(self._make_stats_label(
            f"有效记录：{record_count} 条", sp(15), dp(24), COLOR_TEXT_SECONDARY, halign="left"
        ))
        card.add_widget(summary)
        return card

    def refresh_stats_page(self, reset_period=False):
        """用共享数据重建统计结果；页面控件本身始终只创建一次。"""
        periods = self._get_stats_periods()
        current_period = datetime.now().strftime("%Y" if self.stats_mode == "year" else "%Y-%m")
        selected_period = self.stats_period_spinner.text
        if reset_period or selected_period not in periods:
            selected_period = current_period

        self._updating_stats_period = True
        self.stats_period_spinner.values = periods
        self.stats_period_spinner.text = selected_period
        self._updating_stats_period = False

        self.stats_results.clear_widgets()
        if self.stats_mode == "year":
            records = self.get_records_for_year(selected_period)
            empty_text = "暂无本年度消费记录"
        else:
            records = self.get_records_for_month(selected_period)
            empty_text = "暂无本月消费记录"

        # 收入/支出严格分离：统计、分类、排名只处理支出记录；收入单独汇总
        expense_records = [
            record for record in records
            if isinstance(record, dict) and self._get_record_type(record) == "支出"
        ]
        income_records = [
            record for record in records
            if isinstance(record, dict) and self._get_record_type(record) == "收入"
        ]
        income_total = sum(
            amount for record in income_records
            if (amount := self._get_record_amount(record)) is not None
        )

        category_stats = self.get_category_stats(expense_records)
        total = sum(amount for category, amount in category_stats)
        valid_record_count = sum(1 for record in expense_records if self._get_record_amount(record) is not None)

        if self.income_enabled:
            self.stats_results.add_widget(
                self._build_income_summary_card(selected_period, income_total, total)
            )
        self.stats_results.add_widget(
            self._build_stats_summary_card(selected_period, total, valid_record_count)
        )

        details_card = self._make_card()
        details = BoxLayout(
            orientation="vertical", spacing=dp(12), padding=[CARD_PADDING] * 4, size_hint_y=None
        )
        details.bind(minimum_height=details.setter("height"))
        details.add_widget(self._build_category_stats_section(category_stats, total, empty_text, expense_records, selected_period))
        if self.stats_mode == "year":
            monthly_totals = self.get_monthly_totals_for_year(selected_period)
            details.add_widget(self._build_year_monthly_totals_section(monthly_totals))
        details_card.add_widget(details)
        self.stats_results.add_widget(details_card)

    def show_monthly_stats(self, instance):
        current_month = datetime.now().strftime("%Y-%m")
        self.show_monthly_stats_popup(current_month)

    def show_history_stats(self, instance):
        months = self.get_available_months()
        years = self.get_available_years()

        if not months and not years:
            self.show_popup("提示", "暂无历史记录。")
            return

        content = BoxLayout(orientation="vertical", spacing=dp(12), padding=dp(12))

        type_label = Label(
            text="统计类型",
            font_size=sp(16),
            size_hint_y=None,
            height=dp(26),
            halign="left",
            valign="middle",
            color=COLOR_TEXT
        )
        type_label.bind(size=lambda inst, val: setattr(inst, "text_size", (val[0], val[1])))
        content.add_widget(type_label)

        type_spinner = self._make_spinner(
            "按月统计" if months else "按年统计", ("按月统计", "按年统计")
        )
        content.add_widget(type_spinner)

        period_label = Label(
            text="统计周期",
            font_size=sp(16),
            size_hint_y=None,
            height=dp(26),
            halign="left",
            valign="middle",
            color=COLOR_TEXT
        )
        period_label.bind(size=lambda inst, val: setattr(inst, "text_size", (val[0], val[1])))
        content.add_widget(period_label)

        initial_periods = months if type_spinner.text == "按月统计" else years
        period_spinner = self._make_spinner(
            initial_periods[0] if initial_periods else "", initial_periods
        )
        content.add_widget(period_spinner)

        btn_view = self._make_primary_button("查看统计")
        btn_close = self._make_text_button("关闭")

        popup = self._make_popup("历史统计", content, (0.90, 0.56))

        def update_periods(spinner, text):
            periods = months if text == "按月统计" else years
            period_spinner.values = periods
            period_spinner.text = periods[0] if periods else ""

        def do_view(btn):
            if not period_spinner.text:
                self.show_popup("提示", "暂无历史记录。")
                return
            popup.dismiss()
            if type_spinner.text == "按年统计":
                self.show_stats_for_period("year", period_spinner.text)
            else:
                self.show_stats_for_period("month", period_spinner.text)

        type_spinner.bind(text=update_periods)
        btn_view.bind(on_press=do_view)
        btn_close.bind(on_press=popup.dismiss)

        content.add_widget(btn_view)
        content.add_widget(btn_close)
        popup.open()

    def _get_record_date(self, record):
        if not isinstance(record, dict):
            return None
        try:
            return datetime.strptime(str(record.get("日期", "")), "%Y-%m-%d")
        except Exception:
            return None

    def _get_record_amount(self, record):
        if not isinstance(record, dict):
            return None
        try:
            amount = float(record.get("金额", 0))
            return amount if amount > 0 else None
        except Exception:
            return None

    def get_available_months(self):
        months = set()
        for record in self.records:
            date_obj = self._get_record_date(record)
            if date_obj is not None:
                months.add(date_obj.strftime("%Y-%m"))
        return sorted(months, reverse=True)

    def get_available_years(self):
        years = set()
        for record in self.records:
            date_obj = self._get_record_date(record)
            if date_obj is not None:
                years.add(date_obj.strftime("%Y"))
        return sorted(years, reverse=True)

    def get_records_for_month(self, month_str):
        month_records = []
        for record in self.records:
            date_obj = self._get_record_date(record)
            if date_obj is not None and date_obj.strftime("%Y-%m") == month_str:
                month_records.append(record)
        return month_records

    def get_records_for_year(self, year_str):
        year_records = []
        for record in self.records:
            date_obj = self._get_record_date(record)
            if date_obj is not None and date_obj.strftime("%Y") == year_str:
                year_records.append(record)
        return year_records

    def _get_category_name(self, record):
        """与统计一致的分类名称标准化：空值/空字符串统一归为“未分类”。"""
        return str(record.get("分类", "未分类")).strip() or "未分类"

    def get_category_stats(self, records):
        category_stats = {}
        for record in records:
            amount = self._get_record_amount(record)
            if amount is None:
                continue
            category = self._get_category_name(record)
            category_stats[category] = category_stats.get(category, 0.0) + amount
        return sorted(category_stats.items(), key=lambda x: x[1], reverse=True)

    def get_monthly_totals_for_year(self, year_str):
        monthly_totals = {month: 0.0 for month in range(1, 13)}
        for record in self.get_records_for_year(year_str):
            date_obj = self._get_record_date(record)
            amount = self._get_record_amount(record)
            if date_obj is None or amount is None:
                continue
            monthly_totals[date_obj.month] += amount
        return monthly_totals

    def _make_stats_label(self, text, font_size, height, color, halign="center"):
        label = Label(
            text=text,
            font_size=font_size,
            size_hint_y=None,
            height=height,
            halign=halign,
            valign="middle",
            color=color
        )
        label.bind(size=lambda inst, val: setattr(inst, "text_size", (val[0], val[1])))
        return label

    def _make_category_detail_row(self, category, amount, total, color):
        row = BoxLayout(orientation="horizontal", size_hint_y=None, height=dp(54), spacing=dp(8), padding=(0, dp(4)))

        marker_box = BoxLayout(size_hint=(None, 1), width=dp(18), padding=(0, dp(16), 0, dp(16)))
        marker = Widget(size_hint=(None, None), size=(dp(14), dp(14)))
        with marker.canvas.before:
            Color(*color)
            marker.color_box = RoundedRectangle(radius=[dp(3)] * 4, pos=marker.pos, size=marker.size)

        def update_marker(instance, value):
            instance.color_box.pos = instance.pos
            instance.color_box.size = instance.size

        marker.bind(pos=update_marker, size=update_marker)
        marker_box.add_widget(marker)
        row.add_widget(marker_box)

        category_label = Label(
            text=category,
            font_size=sp(15),
            size_hint=(1, None),
            height=dp(46),
            halign="left",
            valign="middle",
            color=COLOR_TEXT
        )
        category_label.bind(size=lambda inst, val: setattr(inst, "text_size", (val[0], val[1])))
        row.add_widget(category_label)

        percent = amount / total * 100 if total > 0 else 0
        amount_label = Label(
            text=f"{amount:.2f} 元\n{percent:.1f}%",
            font_size=sp(14),
            size_hint=(None, None),
            width=dp(104),
            height=dp(46),
            halign="right",
            valign="middle",
            color=COLOR_TEXT
        )
        amount_label.bind(size=lambda inst, val: setattr(inst, "text_size", (val[0], val[1])))
        row.add_widget(amount_label)
        return row

    def _build_category_stats_section(self, category_stats, total, empty_text, records, selected_period):
        section = BoxLayout(orientation="vertical", spacing=dp(6), size_hint_y=None)
        section.bind(minimum_height=section.setter("height"))

        title_row = BoxLayout(size_hint_y=None, height=dp(32), spacing=dp(8))
        title_label = self._make_stats_label("分类消费", sp(17), dp(32), COLOR_TEXT, halign="left")
        rank_button = self._make_text_button("消费排名", height=dp(32), font_size=sp(15))
        rank_button.size_hint_x = None
        rank_button.width = dp(100)
        rank_button.bind(on_press=lambda instance: self.show_top_consumption_ranking(records, selected_period))
        title_row.add_widget(title_label)
        title_row.add_widget(rank_button)
        section.add_widget(title_row)

        if total > 0:
            chart = CategoryPieChart(
                category_stats,
                size_hint_y=None,
                height=dp(190),
                on_category_press=lambda category: self.show_category_records(
                    category, records, selected_period
                ),
            )
            section.add_widget(chart)
        else:
            section.add_widget(self._make_stats_label(
                empty_text,
                sp(17),
                dp(110),
                COLOR_TEXT_SECONDARY
            ))

        section.add_widget(self._make_stats_label(
            "分类明细",
            sp(17),
            dp(28),
            COLOR_TEXT,
            halign="left"
        ))

        if total > 0:
            for index, (category, amount) in enumerate(category_stats):
                section.add_widget(self._make_category_detail_row(
                    category, amount, total, CATEGORY_CHART_COLORS[index % len(CATEGORY_CHART_COLORS)]
                ))
        else:
            section.add_widget(self._make_stats_label(
                empty_text,
                sp(16),
                dp(52),
                COLOR_TEXT_SECONDARY
            ))
        return section

    def show_category_records(self, category, records, period):
        """显示某分类在当前统计周期内的消费记录；复用 RecordRow 与 show_record_detail。"""
        period_suffix = "年度统计" if self.stats_mode == "year" else "月度统计"
        category_records = [
            record for record in records
            if isinstance(record, dict) and self._get_category_name(record) == category
        ]
        category_records.sort(key=self._get_record_sort_datetime, reverse=True)

        total = 0.0
        valid_count = 0
        for record in category_records:
            amount = self._get_record_amount(record)
            if amount is not None:
                total += amount
                valid_count += 1

        content = BoxLayout(orientation="vertical", spacing=dp(10), padding=dp(10))

        summary = BoxLayout(orientation="vertical", spacing=dp(2), size_hint_y=None, height=dp(76))
        summary.add_widget(self._make_stats_label(
            f"{period}  {period_suffix}", sp(14), dp(24), COLOR_TEXT_SECONDARY, halign="left"
        ))
        summary.add_widget(self._make_stats_label(
            f"总支出：{total:.2f} 元", sp(18), dp(28), COLOR_TEXT, halign="left"
        ))
        summary.add_widget(self._make_stats_label(
            f"记录数量：{valid_count} 条", sp(14), dp(24), COLOR_TEXT_SECONDARY, halign="left"
        ))
        content.add_widget(summary)

        scroll = ScrollView(size_hint=(1, 1), do_scroll_x=False)
        list_box = BoxLayout(orientation="vertical", spacing=dp(8), size_hint_y=None)
        list_box.bind(minimum_height=list_box.setter("height"))
        if category_records:
            for record in category_records:
                list_box.add_widget(RecordRow(record, self.show_record_detail))
        else:
            list_box.add_widget(self._make_stats_label(
                "该分类在本周期内没有有效消费记录", sp(15), dp(120), COLOR_TEXT_SECONDARY
            ))
        scroll.add_widget(list_box)
        content.add_widget(scroll)

        close_btn = self._make_text_button("关闭")
        content.add_widget(close_btn)

        popup = self._make_popup(category, content, (0.92, 0.86))
        close_btn.bind(on_press=popup.dismiss)
        popup.open()

    def get_top_consumption_records(self, records, limit=20):
        """当前周期内单笔消费金额最高的前 limit 条记录。

        只允许有效正数金额参与；排序键为（金额，记录时间）双降序，
        金额相同时记录时间较新的排前面；不修改原始 records。
        """
        valid_records = [
            record for record in records
            if isinstance(record, dict) and self._get_record_amount(record) is not None
        ]
        valid_records.sort(
            key=lambda record: (
                self._get_record_amount(record),
                self._get_record_sort_datetime(record),
            ),
            reverse=True,
        )
        return valid_records[:limit]

    def show_top_consumption_ranking(self, records, period):
        """显示当前统计周期内单笔消费金额最高的前 20 条记录。"""
        period_suffix = "年度统计" if self.stats_mode == "year" else "月度统计"
        ranked = self.get_top_consumption_records(records)

        content = BoxLayout(orientation="vertical", spacing=dp(10), padding=dp(10))

        summary = BoxLayout(orientation="vertical", spacing=dp(2), size_hint_y=None, height=dp(60))
        summary.add_widget(self._make_stats_label(
            f"{period}  {period_suffix}", sp(14), dp(22), COLOR_TEXT_SECONDARY, halign="left"
        ))
        summary.add_widget(self._make_stats_label(
            "当前周期单笔消费前20", sp(16), dp(26), COLOR_TEXT, halign="left"
        ))
        content.add_widget(summary)

        scroll = ScrollView(size_hint=(1, 1), do_scroll_x=False)
        list_box = BoxLayout(orientation="vertical", spacing=dp(8), size_hint_y=None)
        list_box.bind(minimum_height=list_box.setter("height"))
        if ranked:
            for index, record in enumerate(ranked):
                list_box.add_widget(self._make_ranking_record_row(index + 1, record))
        else:
            list_box.add_widget(self._make_stats_label(
                "当前周期暂无有效消费记录", sp(15), dp(120), COLOR_TEXT_SECONDARY
            ))
        scroll.add_widget(list_box)
        content.add_widget(scroll)

        close_btn = self._make_text_button("关闭")
        content.add_widget(close_btn)

        popup = self._make_popup("单笔消费排名", content, (0.92, 0.86))
        close_btn.bind(on_press=popup.dismiss)
        popup.open()

    def _make_ranking_record_row(self, rank, record):
        return RecordRow(record, self.show_record_detail, rank=rank)

    def _make_month_total_row(self, month, amount):
        row = BoxLayout(orientation="horizontal", size_hint_y=None, height=dp(38), spacing=dp(8))
        month_label = Label(
            text=f"{month}月",
            font_size=sp(15),
            size_hint=(0.35, None),
            height=dp(38),
            halign="left",
            valign="middle",
            color=COLOR_TEXT
        )
        month_label.bind(size=lambda inst, val: setattr(inst, "text_size", (val[0], val[1])))
        row.add_widget(month_label)

        amount_label = Label(
            text=f"{amount:.2f}元",
            font_size=sp(15),
            size_hint=(0.65, None),
            height=dp(38),
            halign="right",
            valign="middle",
            color=COLOR_TEXT
        )
        amount_label.bind(size=lambda inst, val: setattr(inst, "text_size", (val[0], val[1])))
        row.add_widget(amount_label)
        return row

    def _build_year_monthly_totals_section(self, monthly_totals):
        section = BoxLayout(orientation="vertical", spacing=dp(4), size_hint_y=None)
        section.bind(minimum_height=section.setter("height"))
        section.add_widget(self._make_stats_label(
            "1月至12月支出明细",
            sp(17),
            dp(30),
            COLOR_TEXT,
            halign="left"
        ))
        for month in range(1, 13):
            section.add_widget(self._make_month_total_row(month, monthly_totals.get(month, 0.0)))
        return section

    def show_monthly_stats_popup(self, month_str):
        self.show_stats_for_period("current_month", month_str)

    def show_stats_for_month(self, month_str):
        self.show_stats_for_period("month", month_str)

    def show_stats_for_period(self, period_type, period_value):
        is_year = period_type == "year"
        is_current_month = period_type == "current_month"

        if is_year:
            records = [
                r for r in self.get_records_for_year(period_value)
                if isinstance(r, dict) and self._get_record_type(r) == "支出"
            ]
            category_stats = self.get_category_stats(records)
            total = sum(amount for category, amount in category_stats)
            monthly_totals = self.get_monthly_totals_for_year(period_value)
            title_text = f"{period_value} 年统计"
            total_text = f"年度总支出：{total:.2f} 元" if total > 0 else f"{period_value} 年没有有效消费记录"
            empty_text = f"{period_value} 年没有有效消费记录"
            popup_title = "年度统计"
        else:
            records = [
                r for r in self.get_records_for_month(period_value)
                if isinstance(r, dict) and self._get_record_type(r) == "支出"
            ]
            category_stats = self.get_category_stats(records)
            total = sum(amount for category, amount in category_stats)
            monthly_totals = None
            title_text = f"{period_value} 本月统计" if is_current_month else f"{period_value} 月统计"
            total_prefix = "本月总支出" if is_current_month else "月度总支出"
            total_text = f"{total_prefix}：{total:.2f} 元" if total > 0 else f"{period_value} 没有有效消费记录"
            empty_text = "本月暂无消费记录" if is_current_month else f"{period_value} 没有有效消费记录"
            popup_title = "本月统计" if is_current_month else "月度统计"

        content = BoxLayout(orientation="vertical", spacing=dp(10), padding=dp(12))
        content.add_widget(self._make_stats_label(
            title_text,
            sp(20),
            dp(34),
            COLOR_TEXT
        ))
        content.add_widget(self._make_stats_label(
            total_text,
            sp(18),
            dp(32),
            COLOR_SUCCESS
        ))

        scroll = ScrollView(size_hint=(1, 1), do_scroll_x=False)
        scroll_content = BoxLayout(orientation="vertical", spacing=dp(10), size_hint_y=None)
        scroll_content.bind(minimum_height=scroll_content.setter("height"))
        scroll_content.add_widget(self._build_category_stats_section(category_stats, total, empty_text, records, period_value))

        if is_year:
            scroll_content.add_widget(self._build_year_monthly_totals_section(monthly_totals))

        scroll.add_widget(scroll_content)
        content.add_widget(scroll)

        btn_close = self._make_text_button("关闭")
        content.add_widget(btn_close)

        popup = self._make_popup(popup_title, content, (0.94, 0.90))
        btn_close.bind(on_press=popup.dismiss)
        popup.open()

    # =========================
    # 分类设置
    # =========================
    def show_categories(self, instance):
        content = BoxLayout(orientation="vertical", spacing=dp(10), padding=dp(10))

        scroll = ScrollView(size_hint=(1, 1))
        self.categories_grid = GridLayout(cols=1, spacing=dp(8), size_hint_y=None)
        self.categories_grid.bind(minimum_height=self.categories_grid.setter("height"))
        scroll.add_widget(self.categories_grid)
        content.add_widget(scroll)

        self.new_category_input = self._make_text_input(hint_text="输入新分类")
        content.add_widget(self.new_category_input)

        type_row = BoxLayout(size_hint_y=None, height=dp(44), spacing=dp(8))
        type_row.add_widget(self._make_section_label("是否为收入", font_size=sp(16), height=dp(44)))
        self.new_category_income_switch = IncomeTypeSwitch(
            active=False, pos_hint={"center_y": 0.5}
        )
        type_row.add_widget(self.new_category_income_switch)
        content.add_widget(type_row)

        add_btn = self._make_primary_button("添加分类")
        add_btn.bind(on_press=self.add_category)
        content.add_widget(add_btn)

        close_btn = self._make_text_button("关闭")
        content.add_widget(close_btn)

        popup = self._make_popup("分类设置", content, (0.9, 0.9))
        self._categories_popup = popup
        close_btn.bind(on_press=popup.dismiss)
        popup.bind(on_dismiss=self._clear_categories_popup_refs)
        self._refresh_categories_grid()
        popup.open()

    def _clear_categories_popup_refs(self, instance):
        if getattr(self, "_categories_popup", None) is instance:
            self._categories_popup = None
            self.categories_grid = None
            self.new_category_input = None
            self.new_category_income_switch = None

    def _refresh_categories_grid(self):
        grid = getattr(self, "categories_grid", None)
        if grid is None:
            return

        grid.clear_widgets()
        for category in self.categories:
            # 整行背景：内边距容纳名称区与 2×2 按钮区，行与行之间由 grid spacing 分隔
            row = BoxLayout(
                size_hint_y=None, height=dp(108), spacing=dp(8),
                padding=[dp(12), dp(10)]
            )
            # ---- 整行柔和圆角背景：确定性映射（crc32），同一分类跨刷新/重启/排序颜色恒定 ----
            with row.canvas.before:
                Color(*_category_bg_color(category))
                row.bg_rect = RoundedRectangle(
                    pos=row.pos, size=row.size, radius=[dp(12)] * 4
                )

            def update_bg(instance, value):
                instance.bg_rect.pos = instance.pos
                instance.bg_rect.size = instance.size

            row.bind(pos=update_bg, size=update_bg)

            # ---- 分类名称：占满左侧剩余宽高，左上对齐（halign=left + valign=top），
            #     三级适配（正常字号单行 → 轻微缩小单行 → 两行） ----
            name_label = Label(
                text=category, color=COLOR_TEXT, font_size=sp(17),
                halign="left", valign="top", shorten=False, max_lines=1,
                size_hint=(1, 1)
            )
            name_label._fit_lock = False
            name_label._two_line = False

            def settle_name(inst, avail):
                if inst._fit_lock:
                    Clock.schedule_once(lambda dt: settle_name(inst, avail), 0.05)
                    return
                inst._fit_lock = True
                try:
                    if inst._two_line:
                        # 已锁定两行模式：保持两行，不再回退单行
                        inst.max_lines = 2
                        inst.shorten = False
                        inst.text_size = (avail, inst.height)
                        return
                    inst.texture_update()  # 强制按当前字号重建纹理，否则 texture_size 返回旧值
                    tex_w = inst.texture_size[0]
                    if tex_w > avail:
                        if inst.font_size > sp(13):
                            # 第二级：按实际渲染宽度逐步轻微缩小字号
                            inst.font_size = max(sp(13), inst.font_size - sp(1))
                            Clock.schedule_once(lambda dt: settle_name(inst, avail), 0.05)
                            return
                        # 第三级：达到最小字号仍超宽 → 两行
                        inst._two_line = True
                        inst.max_lines = 2
                        inst.shorten = False
                        inst.text_size = (avail, inst.height)
                        return
                    # 第一级：正常字号单行
                    inst.max_lines = 1
                    inst.text_size = (avail, inst.height)
                finally:
                    inst._fit_lock = False

            def fit_name(inst, val):
                if inst._fit_lock:
                    return
                if inst.width <= 0:
                    return
                inst._fit_lock = True
                try:
                    if inst._two_line:
                        inst.max_lines = 2
                        inst.shorten = False
                        inst.text_size = (inst.width, inst.height)
                        return
                    inst.text_size = (None, None)  # 完整渲染，测量真实宽度
                    Clock.schedule_once(lambda dt: settle_name(inst, inst.width), 0.05)
                finally:
                    inst._fit_lock = False

            # 只绑定 size：text_size 的赋值会改变 texture，若再绑定 texture_size
            # 会与 settle_name 形成 0.05s 震荡循环（text_size 被反复重置）。
            # 收敛由 settle_name 内部的延迟重测驱动。
            name_label.bind(size=fit_name)
            row.add_widget(name_label)

            # ---- 右侧 2×2 操作区（上移/下移 左列，类型/删除 右列） ----
            btn_area = GridLayout(
                cols=2, spacing=dp(4), size_hint=(None, None),
                width=dp(132), height=dp(88)
            )
            move_up_btn = self._make_secondary_button("上移", font_size=sp(12))
            move_up_btn.size_hint_y = 1
            move_up_btn.bind(on_press=lambda btn, cat=category: self.move_category(cat, -1))
            btn_area.add_widget(move_up_btn)

            # 分类类型切换：只影响以后新记记录的类型，不改历史记录
            if self._is_income_category(category):
                type_btn = self._make_button(
                    "收入", COLOR_SUCCESS_LIGHT, COLOR_SUCCESS, font_size=sp(12)
                )
            else:
                type_btn = self._make_button(
                    "支出", COLOR_DANGER_LIGHT, COLOR_DANGER, font_size=sp(12)
                )
            type_btn.size_hint_y = 1
            type_btn.bind(on_press=lambda btn, cat=category: self.toggle_category_type(cat))
            btn_area.add_widget(type_btn)

            move_down_btn = self._make_secondary_button("下移", font_size=sp(12))
            move_down_btn.size_hint_y = 1
            move_down_btn.bind(on_press=lambda btn, cat=category: self.move_category(cat, 1))
            btn_area.add_widget(move_down_btn)

            delete_btn = self._make_button(
                "删除", (0.945, 0.949, 0.957, 1), COLOR_DANGER, font_size=sp(12)
            )
            delete_btn.size_hint_y = 1
            delete_btn.bind(on_press=lambda btn, cat=category: self.confirm_delete_category(cat))
            btn_area.add_widget(delete_btn)

            row.add_widget(btn_area)
            grid.add_widget(row)

    def _sync_category_spinner(self, removed_category=None):
        # 收入功能关闭时隐藏收入分类，避免误记收入；开启后恢复全部可见
        if self.income_enabled:
            visible_categories = list(self.categories)
        else:
            visible_categories = [
                c for c in self.categories if c not in self.income_categories
            ]
        self.category_spinner.values = visible_categories
        if visible_categories and (self.category_spinner.text not in visible_categories or self.category_spinner.text == removed_category):
            self.category_spinner.text = visible_categories[0]

    def add_category(self, instance):
        new_category = self.new_category_input.text.strip()
        if not new_category:
            self.show_popup("提示", "请输入分类名称。")
            return

        if new_category in self.categories:
            self.show_popup("提示", "该分类已存在。")
            return

        self.categories.append(new_category)
        if getattr(self, "new_category_income_switch", None) is not None and self.new_category_income_switch.active:
            self.income_categories.add(new_category)
            self._save_settings()
        self._sync_category_spinner()
        self.save_data()
        self.new_category_input.text = ""
        if getattr(self, "new_category_income_switch", None) is not None:
            self.new_category_income_switch.active = False
        self._refresh_categories_grid()
        self.refresh_settings_page()
        self.show_popup("成功", f"已添加分类：{new_category}")

    def confirm_delete_category(self, category):
        message = f"确定要删除分类“{category}”吗？\n历史账单中的分类记录不会被修改。"
        self.show_confirm_popup("确认删除分类", message, lambda btn: self.delete_category(category))

    def delete_category(self, category):
        if category not in self.categories:
            self._refresh_categories_grid()
            return

        if len(self.categories) <= 1:
            self.show_popup("提示", "至少保留一个分类。")
            self._refresh_categories_grid()
            return

        self.categories.remove(category)
        if category in self.income_categories:
            self.income_categories.discard(category)
            self._save_settings()
        self._sync_category_spinner(removed_category=category)
        self.save_data()
        self._refresh_categories_grid()
        self.refresh_settings_page()
        self.show_popup("成功", f"已删除分类：{category}")

    def move_category(self, category, offset):
        """上移(offset=-1)或下移(offset=1)一个分类。

        只改变 self.categories 的顺序，不修改历史账单中的分类字段；
        边界外操作（第一项上移、最后一项下移、分类不存在）安全返回。
        """
        if offset == 0 or category not in self.categories:
            return
        index = self.categories.index(category)
        target = index + offset
        if target < 0 or target >= len(self.categories):
            return
        self.categories[index], self.categories[target] = self.categories[target], self.categories[index]
        self._sync_category_spinner()
        self.save_data()
        self._refresh_categories_grid()

    def toggle_category_type(self, category):
        """切换分类为收入/支出分类；只影响以后新记记录的类型，不修改历史记录。"""
        if self._is_income_category(category):
            self.income_categories.discard(category)
        else:
            self.income_categories.add(category)
        self._save_settings()
        self._sync_category_spinner()
        self._refresh_categories_grid()

    # =========================
    # 查看记录
    # =========================
    def show_records(self, instance):
        if not self.records:
            self.show_popup("提示", "暂无记录。")
            return

        self.sort_records()

        content = BoxLayout(orientation="vertical", spacing=dp(10), padding=dp(10))
        scroll = ScrollView(size_hint=(1, 1))
        grid = GridLayout(cols=1, spacing=dp(8), size_hint_y=None)
        grid.bind(minimum_height=grid.setter("height"))

        display_records = [record for record in self.records if isinstance(record, dict)][:50]

        for record in display_records:
            note = str(record.get("姓名/备注", ""))
            category = str(record.get("分类", ""))
            try:
                amount_text = f"{float(record.get('金额', 0)):.2f}"
            except (TypeError, ValueError):
                amount_text = str(record.get("金额", ""))
            date_str = str(record.get("日期", ""))

            text = f"{date_str}  {category}\n{amount_text}元  {note}"
            row = Label(
                text=text,
                font_size=sp(16),
                size_hint_y=None,
                height=dp(62),
                halign="left",
                valign="middle",
                color=COLOR_TEXT
            )
            row.bind(size=lambda inst, val: setattr(inst, "text_size", (val[0] - dp(10), None)))
            grid.add_widget(row)

        scroll.add_widget(grid)
        content.add_widget(scroll)

        close_btn = self._make_text_button("关闭")
        content.add_widget(close_btn)

        popup = self._make_popup("查看记录（最近50条）", content, (0.92, 0.9))
        close_btn.bind(on_press=popup.dismiss)
        popup.open()

    # =========================
    # 删除记录
    # =========================
    def delete_records(self, instance):
        if not self.records:
            self.show_popup("提示", "暂无记录可删除。")
            return

        self.sort_records()

        content = BoxLayout(orientation="vertical", spacing=dp(10), padding=dp(10))
        scroll = ScrollView(size_hint=(1, 1))
        grid = GridLayout(cols=1, spacing=dp(8), size_hint_y=None)
        grid.bind(minimum_height=grid.setter("height"))

        self._delete_records_grid = grid

        scroll.add_widget(grid)
        content.add_widget(scroll)

        clear_btn = self._make_danger_button("清空所有记录")
        clear_btn.bind(on_press=self.clear_all_records)
        content.add_widget(clear_btn)

        close_btn = self._make_text_button("关闭")
        content.add_widget(close_btn)

        popup = self._make_popup("删除记录（最近20条）", content, (0.92, 0.9))
        self._delete_records_popup = popup
        close_btn.bind(on_press=popup.dismiss)
        popup.bind(on_dismiss=self._clear_delete_records_popup)
        self._rebuild_delete_records_grid()
        popup.open()

    def _rebuild_delete_records_grid(self):
        """根据当前记录重建删除窗口，始终只展示最近 20 条。"""
        grid = getattr(self, "_delete_records_grid", None)
        if grid is None:
            return

        grid.clear_widgets()
        display_records = [record for record in self.records if isinstance(record, dict)][:20]
        for record in display_records:
            row = BoxLayout(size_hint_y=None, height=dp(68), spacing=dp(8))

            try:
                amount_text = f"{float(record.get('金额', 0)):.2f}"
            except (TypeError, ValueError):
                amount_text = str(record.get("金额", ""))
            record_text = (
                f"{record.get('日期', '')} {record.get('分类', '')}\n"
                f"{amount_text}元 {str(record.get('姓名/备注', ''))[:14]}"
            )

            info_label = Label(
                text=record_text,
                font_size=sp(15),
                size_hint=(0.72, 1),
                halign="left",
                valign="middle",
                color=COLOR_TEXT
            )
            info_label.bind(size=lambda inst, val: setattr(inst, "text_size", (val[0] - dp(6), None)))
            row.add_widget(info_label)

            delete_btn = self._make_danger_button("删除", font_size=sp(16))
            delete_btn.size_hint = (0.28, 1)
            delete_btn.bind(on_press=lambda btn, item=record: self.delete_single_record(item))
            row.add_widget(delete_btn)

            grid.add_widget(row)

    def _clear_delete_records_popup(self, instance):
        if getattr(self, "_delete_records_popup", None) is instance:
            self._delete_records_popup = None
            self._delete_records_grid = None

    def delete_single_record(self, record):
        # 使用对象身份而非旧索引，避免窗口内连续删除时删错记录。
        for index, current_record in enumerate(self.records):
            if current_record is record:
                del self.records[index]
                self.save_data()
                self.update_monthly_expense()
                if self.records:
                    self._rebuild_delete_records_grid()
                else:
                    popup = getattr(self, "_delete_records_popup", None)
                    if popup is not None:
                        popup.dismiss()
                return

    def clear_all_records(self, instance):
        def do_clear(btn):
            self.records = []
            self.save_data()
            self.update_monthly_expense()
            self.refresh_settings_page()
            self.records_page_dirty = True
            self.income_page_dirty = True
            self.stats_page_dirty = True
            popup = getattr(self, "_delete_records_popup", None)
            if popup is not None:
                popup.dismiss()
            self.show_popup("成功", "所有记录已清空。")

        self.show_confirm_popup("确认清空", "确定要清空所有记录吗？此操作不可撤销。", do_clear)

    # =========================
    # 导出
    # =========================
    def export_data(self, instance):
        content = BoxLayout(orientation="vertical", spacing=dp(10), padding=dp(12))

        btn_xlsx = self._make_secondary_button("导出 Excel：账单表格")
        btn_csv = self._make_secondary_button("导出 CSV：账单表格")
        btn_json = self._make_secondary_button("导出完整备份 JSON：账单和分类")
        btn_close = self._make_text_button("关闭")

        popup = self._make_popup("导出数据", content, (0.86, 0.54))

        btn_xlsx.bind(on_press=lambda btn: self.export_to_excel(popup))
        btn_csv.bind(on_press=lambda btn: self.export_to_csv(popup))
        btn_json.bind(on_press=lambda btn: self.export_to_json(popup))
        btn_close.bind(on_press=popup.dismiss)

        content.add_widget(btn_xlsx)
        content.add_widget(btn_csv)
        content.add_widget(btn_json)
        content.add_widget(btn_close)

        popup.open()

    def _is_android(self):
        return platform == "android"

    def _get_export_fieldnames(self):
        return ["记录ID", "记录类型", "姓名/备注", "分类", "金额", "日期", "记录时间"]

    def _get_export_timestamp(self):
        return datetime.now().strftime("%Y%m%d_%H%M%S")

    def _format_exception(self, error):
        return f"{type(error).__name__}: {str(error)}"

    def _create_excel_temp_file(self, timestamp):
        temp_path = os.path.join(self.storage_dir, f"export_temp_{timestamp}.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "记账记录"
        ws.append(self._get_export_fieldnames())

        for record in self.records:
            if not isinstance(record, dict):
                continue
            ws.append([
                record.get("记录ID", ""),
                record.get("记录类型", "支出"),
                record.get("姓名/备注", ""),
                record.get("分类", ""),
                record.get("金额", ""),
                record.get("日期", ""),
                record.get("记录时间", "")
            ])

        wb.save(temp_path)
        return temp_path

    def _create_csv_temp_file(self, timestamp):
        temp_path = os.path.join(self.storage_dir, f"export_temp_{timestamp}.csv")
        fieldnames = self._get_export_fieldnames()
        with open(temp_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for record in self.records:
                if not isinstance(record, dict):
                    continue
                writer.writerow({
                    "记录ID": record.get("记录ID", ""),
                    "记录类型": record.get("记录类型", "支出"),
                    "姓名/备注": record.get("姓名/备注", ""),
                    "分类": record.get("分类", ""),
                    "金额": record.get("金额", ""),
                    "日期": record.get("日期", ""),
                    "记录时间": record.get("记录时间", "")
                })
        return temp_path

    def _create_json_temp_file(self, timestamp):
        temp_path = os.path.join(self.storage_dir, f"export_temp_{timestamp}.json")
        backup_data = {
            "records": self.records,
            "categories": self.categories
        }
        with open(temp_path, "w", encoding="utf-8") as f:
            json.dump(backup_data, f, ensure_ascii=False, indent=2)
        return temp_path

    def _create_export_temp_file(self, export_type, timestamp):
        if export_type == "excel":
            return self._create_excel_temp_file(timestamp)
        if export_type == "csv":
            return self._create_csv_temp_file(timestamp)
        if export_type == "json":
            return self._create_json_temp_file(timestamp)
        raise ValueError(f"未知导出类型：{export_type}")

    def _cleanup_export_temp_file(self, temp_path):
        if not temp_path:
            return None
        try:
            if os.path.exists(temp_path):
                os.remove(temp_path)
        except Exception as e:
            return self._format_exception(e)
        return None

    def _get_android_export_config(self, export_type, timestamp):
        configs = {
            "excel": {
                "request_code": 5101,
                "extension": "xlsx",
                "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "success_message": "Excel 导出成功"
            },
            "csv": {
                "request_code": 5102,
                "extension": "csv",
                "mime_type": "text/csv",
                "success_message": "CSV 导出成功"
            },
            "json": {
                "request_code": 5103,
                "extension": "json",
                "mime_type": "application/json",
                "success_message": "完整备份 JSON 导出成功"
            }
        }
        config = configs[export_type].copy()
        if export_type == "json":
            config["filename"] = f"个人记账完整备份_{timestamp}.json"
        else:
            config["filename"] = f"记账记录_{timestamp}.{config['extension']}"
        return config

    def _ensure_android_export_binding(self):
        if not self._is_android() or self._android_export_bound:
            return
        if platform == "android":
            from android import activity
            activity.bind(on_activity_result=self._on_android_activity_result)
            self._android_export_bound = True

    def _start_android_document_export(self, export_type, popup=None):
        if export_type != "json" and not self._has_exportable_records():
            self.show_popup("提示", "暂无记录可导出。")
            return

        if self._pending_android_export is not None:
            self.show_popup("提示", "已有导出任务正在等待系统保存界面返回，请稍后再试。")
            return

        temp_path = None
        try:
            self.sort_records()
            timestamp = self._get_export_timestamp()
            config = self._get_android_export_config(export_type, timestamp)
            temp_path = self._create_export_temp_file(export_type, timestamp)
            self._ensure_android_export_binding()

            if platform == "android":
                from jnius import autoclass
                Intent = autoclass("android.content.Intent")
                PythonActivity = autoclass("org.kivy.android.PythonActivity")

                intent = Intent(Intent.ACTION_CREATE_DOCUMENT)
                intent.addCategory(Intent.CATEGORY_OPENABLE)
                intent.setType(config["mime_type"])
                intent.putExtra(Intent.EXTRA_TITLE, config["filename"])

                self._pending_android_export = {
                    "type": export_type,
                    "request_code": config["request_code"],
                    "temp_path": temp_path,
                    "filename": config["filename"],
                    "success_message": config["success_message"]
                }

                if popup:
                    popup.dismiss()
                PythonActivity.mActivity.startActivityForResult(intent, config["request_code"])
        except Exception as e:
            cleanup_warning = self._cleanup_export_temp_file(temp_path)
            self._pending_android_export = None
            message = f"导出失败：\n{self._format_exception(e)}"
            if cleanup_warning:
                message += f"\n临时文件清理警告：{cleanup_warning}"
            self.show_popup("错误", message)

    def _on_android_activity_result(self, request_code, result_code, intent):
        Clock.schedule_once(
            lambda dt: self._dispatch_android_activity_result(request_code, result_code, intent),
            0
        )

    def _dispatch_android_activity_result(self, request_code, result_code, intent):
        if request_code == 5201:
            self._handle_android_import_result(result_code, intent)
            return
        self._handle_android_activity_result(request_code, result_code, intent)

    def _handle_android_activity_result(self, request_code, result_code, intent):
        pending = self._pending_android_export
        if not pending or request_code != pending.get("request_code"):
            return

        temp_path = pending.get("temp_path")
        try:
            if platform == "android":
                from jnius import autoclass
                Activity = autoclass("android.app.Activity")

                if result_code != Activity.RESULT_OK:
                    return

                if intent is None:
                    self.show_popup("错误", "导出失败：\nAndroid 系统保存界面未返回数据。")
                    return

                uri = intent.getData()
                if uri is None:
                    self.show_popup("错误", "导出失败：\nAndroid 系统保存界面未返回文件 URI。")
                    return

                self._write_file_to_content_uri(temp_path, uri)
                self.show_popup("成功", pending.get("success_message", "导出成功"))
        except Exception as e:
            self.show_popup("错误", f"导出失败：\n{self._format_exception(e)}")
        finally:
            cleanup_warning = self._cleanup_export_temp_file(temp_path)
            self._pending_android_export = None
            if cleanup_warning:
                print(f"导出临时文件清理失败：{cleanup_warning}")

    def _write_file_to_content_uri(self, source_path, uri):
        if platform != "android":
            raise RuntimeError("Content URI 写入仅支持 Android 平台。")
        if not source_path or not os.path.exists(source_path):
            raise FileNotFoundError(f"临时导出文件不存在：{source_path}")

        from jnius import autoclass
        PythonActivity = autoclass("org.kivy.android.PythonActivity")
        resolver = PythonActivity.mActivity.getContentResolver()
        output_stream = resolver.openOutputStream(uri)
        if output_stream is None:
            raise OSError("ContentResolver.openOutputStream 返回 None。")

        try:
            with open(source_path, "rb") as source_file:
                while True:
                    chunk = source_file.read(64 * 1024)
                    if not chunk:
                        break
                    output_stream.write(chunk)
            output_stream.flush()
        finally:
            output_stream.close()

    def export_to_excel(self, popup=None):
        if not self._has_exportable_records():
            self.show_popup("提示", "暂无记录可导出。")
            return
        if self._is_android():
            self._start_android_document_export("excel", popup)
            return
        try:
            self.sort_records()
            export_dir = self.get_export_dir()
            timestamp = self._get_export_timestamp()
            filename = f"记账记录_{timestamp}.xlsx"
            file_path = os.path.join(export_dir, filename)
            temp_path = self._create_excel_temp_file(timestamp)
            try:
                os.replace(temp_path, file_path)
            except Exception:
                self._cleanup_export_temp_file(temp_path)
                raise

            if popup:
                popup.dismiss()
            self.show_popup("成功", f"数据已导出为：\n{file_path}")
        except Exception as e:
            self.show_popup("错误", f"导出失败：\n{self._format_exception(e)}")

    def export_to_csv(self, popup=None):
        if not self._has_exportable_records():
            self.show_popup("提示", "暂无记录可导出。")
            return
        if self._is_android():
            self._start_android_document_export("csv", popup)
            return
        try:
            self.sort_records()
            export_dir = self.get_export_dir()
            timestamp = self._get_export_timestamp()
            filename = f"记账记录_{timestamp}.csv"
            file_path = os.path.join(export_dir, filename)
            temp_path = self._create_csv_temp_file(timestamp)
            try:
                os.replace(temp_path, file_path)
            except Exception:
                self._cleanup_export_temp_file(temp_path)
                raise

            if popup:
                popup.dismiss()
            self.show_popup("成功", f"数据已导出为：\n{file_path}")
        except Exception as e:
            self.show_popup("错误", f"导出失败：\n{self._format_exception(e)}")

    def export_to_json(self, popup=None):
        if self._is_android():
            self._start_android_document_export("json", popup)
            return
        try:
            self.sort_records()
            export_dir = self.get_export_dir()
            timestamp = self._get_export_timestamp()
            filename = f"个人记账完整备份_{timestamp}.json"
            file_path = os.path.join(export_dir, filename)
            temp_path = self._create_json_temp_file(timestamp)
            try:
                os.replace(temp_path, file_path)
            except Exception:
                self._cleanup_export_temp_file(temp_path)
                raise

            if popup:
                popup.dismiss()
            self.show_popup("成功", f"数据已导出为：\n{file_path}")
        except Exception as e:
            self.show_popup("错误", f"导出失败：\n{self._format_exception(e)}")

    # =========================
    # 导入
    # =========================
    def import_data_popup(self, instance):
        if self._is_android():
            self._start_android_document_import()
            return

        chooser = FileChooserListView(
            path=self.get_default_import_dir(),
            filters=["*.json", "*.csv", "*.xlsx"],
            size_hint=(1, 1)
        )

        content = BoxLayout(orientation="vertical", spacing=dp(10), padding=dp(10))
        content.add_widget(chooser)

        btn_box = BoxLayout(size_hint_y=None, height=dp(44), spacing=dp(10))
        btn_import = self._make_primary_button("导入")
        btn_import.size_hint_y = 1
        btn_cancel = self._make_text_button("取消")
        btn_cancel.size_hint_y = 1
        btn_box.add_widget(btn_import)
        btn_box.add_widget(btn_cancel)
        content.add_widget(btn_box)

        popup = self._make_popup("选择要导入的数据文件", content, (0.94, 0.92))

        def do_import(btn):
            if not chooser.selection:
                self.show_popup("提示", "请先选择一个文件。")
                return
            file_path = chooser.selection[0]
            popup.dismiss()
            self.import_file(file_path)

        btn_import.bind(on_press=do_import)
        btn_cancel.bind(on_press=popup.dismiss)

        popup.open()

    def _start_android_document_import(self):
        if self._pending_android_import:
            self.show_popup("提示", "文件选择器已打开，请先完成或取消当前选择。")
            return

        try:
            self._ensure_android_export_binding()
            if platform == "android":
                from jnius import autoclass
                Intent = autoclass("android.content.Intent")
                PythonActivity = autoclass("org.kivy.android.PythonActivity")

                intent = Intent(Intent.ACTION_OPEN_DOCUMENT)
                intent.addCategory(Intent.CATEGORY_OPENABLE)
                intent.setType("*/*")
                intent.putExtra(Intent.EXTRA_ALLOW_MULTIPLE, False)
                intent.putExtra(Intent.EXTRA_MIME_TYPES, [
                    "application/json",
                    "text/json",
                    "text/csv",
                    "application/csv",
                    "text/comma-separated-values",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "application/octet-stream",
                ])
                self._pending_android_import = True
                PythonActivity.mActivity.startActivityForResult(intent, 5201)
        except Exception as e:
            self._pending_android_import = False
            self.show_popup("错误", f"无法打开系统文件选择器：\n{self._format_exception(e)}")

    def _handle_android_import_result(self, result_code, intent):
        if not self._pending_android_import:
            return

        temp_path = None
        try:
            if platform == "android":
                from jnius import autoclass
                Activity = autoclass("android.app.Activity")

                if result_code != Activity.RESULT_OK:
                    return
                if intent is None:
                    raise ValueError("Android 系统文件选择器未返回数据。")

                uri = intent.getData()
                if uri is None:
                    raise ValueError("Android 系统文件选择器未返回文件 URI。")

                filename, mime_type = self._get_android_document_info(uri)
                extension = self._resolve_import_extension(filename, mime_type)
                if extension is None:
                    self.show_popup("错误", "无法识别文件格式，请选择 JSON、CSV 或 XLSX 文件。")
                    return

                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
                temp_path = os.path.join(self.storage_dir, f"import_temp_{timestamp}.{extension}")
                self._copy_content_uri_to_file(uri, temp_path)
                self.import_file(temp_path)
        except Exception as e:
            self.show_popup("导入失败", f"读取所选文件失败：\n{self._format_exception(e)}")
        finally:
            self._pending_android_import = False
            cleanup_warning = self._cleanup_export_temp_file(temp_path)
            if cleanup_warning:
                print(f"导入临时文件清理失败：{cleanup_warning}")

    def _get_android_document_info(self, uri):
        from jnius import autoclass
        PythonActivity = autoclass("org.kivy.android.PythonActivity")
        OpenableColumns = autoclass("android.provider.OpenableColumns")
        resolver = PythonActivity.mActivity.getContentResolver()
        mime_type = resolver.getType(uri)
        filename = None
        cursor = None
        try:
            cursor = resolver.query(uri, None, None, None, None)
            if cursor is not None and cursor.moveToFirst():
                name_index = cursor.getColumnIndex(OpenableColumns.DISPLAY_NAME)
                if name_index >= 0 and not cursor.isNull(name_index):
                    filename = str(cursor.getString(name_index))
        except Exception as e:
            print(f"读取所选文件名称失败：{self._format_exception(e)}")
        finally:
            if cursor is not None:
                cursor.close()
        return filename, str(mime_type) if mime_type else None

    def _resolve_import_extension(self, filename, mime_type):
        supported_extensions = {".json": "json", ".csv": "csv", ".xlsx": "xlsx"}
        if filename:
            extension = os.path.splitext(filename)[1].lower()
            if extension in supported_extensions:
                return supported_extensions[extension]

        mime_extensions = {
            "application/json": "json",
            "text/json": "json",
            "text/csv": "csv",
            "application/csv": "csv",
            "text/comma-separated-values": "csv",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
        }
        return mime_extensions.get((mime_type or "").lower())

    def _copy_content_uri_to_file(self, uri, target_path):
        if platform != "android":
            raise RuntimeError("Content URI 读取仅支持 Android 平台。")

        from jnius import autoclass
        PythonActivity = autoclass("org.kivy.android.PythonActivity")
        resolver = PythonActivity.mActivity.getContentResolver()
        input_stream = resolver.openInputStream(uri)
        if input_stream is None:
            raise OSError("ContentResolver.openInputStream 返回 None。")

        try:
            buffer = bytearray(64 * 1024)
            with open(target_path, "wb") as target_file:
                while True:
                    count = input_stream.read(buffer)
                    if count == -1:
                        break
                    if count > 0:
                        target_file.write(buffer[:count])
        finally:
            input_stream.close()

    def import_file(self, file_path):
        try:
            imported_records = []
            imported_categories = None
            import_kind = "records"

            if file_path.lower().endswith(".json"):
                with open(file_path, "r", encoding="utf-8") as f:
                    json_data = json.load(f)

                if isinstance(json_data, dict):
                    if "records" not in json_data and "categories" not in json_data:
                        self.show_popup("导入失败", "文件内容格式不正确。")
                        return
                    imported_records = json_data.get("records", [])
                    imported_categories = json_data.get("categories", [])
                    if not isinstance(imported_records, list) or not isinstance(imported_categories, list):
                        self.show_popup("导入失败", "文件内容格式不正确。")
                        return
                    import_kind = "backup"
                elif isinstance(json_data, list) and all(isinstance(item, str) for item in json_data):
                    imported_categories = json_data
                    import_kind = "categories"
                else:
                    imported_records = json_data

            elif file_path.lower().endswith(".csv"):
                with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
                    reader = csv.DictReader(f)
                    imported_records = list(reader)

            elif file_path.lower().endswith(".xlsx"):
                wb = load_workbook(file_path, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    self.show_popup("导入失败", "Excel 文件为空。")
                    return

                headers = [str(h).strip() if h is not None else "" for h in rows[0]]
                for row in rows[1:]:
                    item = {}
                    for i, header in enumerate(headers):
                        if header:
                            item[header] = row[i] if i < len(row) else ""
                    imported_records.append(item)

            else:
                self.show_popup("错误", "不支持的文件格式。")
                return

            if not isinstance(imported_records, list):
                self.show_popup("导入失败", "文件内容格式不正确。")
                return

            # 幂等判定只使用来源唯一标识（记录ID）：
            # 相同 ID -> 同一来源记录，跳过；不同 ID（或没有 ID）-> 视为不同记录，全部导入。
            # 不再按“备注+分类+金额+日期”内容去重，避免把真实不同的消费误判为重复。
            existing_ids = set()
            for record in self.records:
                if isinstance(record, dict):
                    rid = record.get("记录ID")
                    if isinstance(rid, str) and rid.strip():
                        existing_ids.add(rid.strip())

            valid_records = []
            new_categories = set()
            duplicate_count = 0
            seen_ids_in_file = set()

            for record in imported_records:
                if not isinstance(record, dict):
                    continue

                note = record.get("姓名/备注", record.get("备注", ""))
                category = record.get("分类", "")
                amount = record.get("金额", "")
                date_str = record.get("日期", "")
                record_time = record.get("记录时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

                if str(note).strip() == "" or str(category).strip() == "" or str(date_str).strip() == "":
                    continue

                try:
                    amount = round(float(amount), 2)
                    if amount <= 0:
                        continue
                except Exception:
                    continue

                try:
                    if isinstance(date_str, datetime):
                        date_str = date_str.strftime("%Y-%m-%d")
                    else:
                        date_str = datetime.strptime(str(date_str)[:10], "%Y-%m-%d").strftime("%Y-%m-%d")
                except Exception:
                    continue

                clean_note = str(note).strip()
                clean_category = str(category).strip()
                rid = record.get("记录ID", record.get("record_id", ""))
                if isinstance(rid, str):
                    rid = rid.strip()

                # 只有来源带有可靠唯一标识时才做幂等跳过；无 ID 的记录一律导入
                if rid and (rid in existing_ids or rid in seen_ids_in_file):
                    duplicate_count += 1
                    continue

                record_type = str(record.get("记录类型", "支出")).strip()
                if record_type not in ("收入", "支出"):
                    record_type = "支出"

                clean_record = {
                    "记录ID": rid if rid else uuid.uuid4().hex,
                    "记录类型": record_type,
                    "姓名/备注": clean_note,
                    "分类": clean_category,
                    "金额": amount,
                    "日期": date_str,
                    "记录时间": str(record_time)
                }

                valid_records.append(clean_record)
                new_categories.add(clean_category)
                if rid:
                    existing_ids.add(rid)
                    seen_ids_in_file.add(rid)

            if import_kind == "records" and not valid_records and duplicate_count > 0:
                self.show_popup("导入完成", f"没有新增记录。\n检测到 {duplicate_count} 条已导入过的记录（按记录ID判定），已自动跳过。")
                return

            if import_kind == "records" and not valid_records:
                self.show_popup("导入失败", "文件中没有找到可导入的有效记录。")
                return

            self.records.extend(valid_records)
            if valid_records:
                self.sort_records()

            added_category_count = 0
            duplicate_category_count = 0
            if imported_categories is not None:
                for category in imported_categories:
                    if not isinstance(category, str):
                        continue
                    clean_category = category.strip()
                    if not clean_category:
                        continue
                    if clean_category in self.categories:
                        duplicate_category_count += 1
                        continue
                    self.categories.append(clean_category)
                    added_category_count += 1

            for cat in sorted(new_categories):
                if cat and cat not in self.categories:
                    self.categories.append(cat)

            self.category_spinner.values = self.categories
            if self.category_spinner.text not in self.categories and self.categories:
                self.category_spinner.text = self.categories[0]

            self.save_data()
            self.update_monthly_expense()
            self.refresh_settings_page()
            self.records_page_dirty = True
            self.income_page_dirty = True
            self.stats_page_dirty = True

            if import_kind == "categories":
                self.show_popup(
                    "导入完成",
                    f"新增分类 {added_category_count} 个。\n重复分类 {duplicate_category_count} 个。"
                )
            elif import_kind == "backup":
                self.show_popup(
                    "导入完成",
                    f"成功导入 {len(valid_records)} 条记录。\n"
                    f"自动跳过 {duplicate_count} 条已导入过的记录。\n"
                    f"新增分类 {added_category_count} 个。\n"
                    f"重复分类 {duplicate_category_count} 个。"
                )
            else:
                self.show_popup(
                    "导入成功",
                    f"成功导入 {len(valid_records)} 条记录。\n自动跳过 {duplicate_count} 条已导入过的记录。"
                )

        except Exception as e:
            self.show_popup("导入失败", f"发生错误：\n{str(e)}")

    # =========================
    # 通用弹窗
    # =========================
    def show_popup(self, title, message):
        content = BoxLayout(orientation="vertical", spacing=CARD_SPACING, padding=CARD_PADDING)

        if "成功" in title or "导入完成" in title:
            status_color = COLOR_SUCCESS
            status_background = COLOR_SUCCESS_LIGHT
        elif "错误" in title or "失败" in title:
            status_color = COLOR_DANGER
            status_background = COLOR_DANGER_LIGHT
        else:
            status_color = COLOR_PRIMARY
            status_background = COLOR_PRIMARY_LIGHT

        self._add_rounded_background(content, status_background, INPUT_RADIUS)

        msg = Label(
            text=message,
            font_size=sp(17),
            halign="center",
            valign="middle",
            color=status_color
        )
        msg.bind(size=lambda inst, val: setattr(inst, "text_size", (val[0] - dp(8), None)))
        content.add_widget(msg)

        btn = self._make_button("确定", status_color, COLOR_WHITE)
        content.add_widget(btn)

        popup = self._make_popup(title, content, (0.88, 0.42))
        btn.bind(on_press=popup.dismiss)
        popup.open()

    def show_confirm_popup(self, title, message, confirm_callback):
        content = BoxLayout(orientation="vertical", spacing=CARD_SPACING, padding=CARD_PADDING)

        msg = Label(
            text=message,
            font_size=sp(17),
            halign="center",
            valign="middle",
            color=COLOR_TEXT
        )
        msg.bind(size=lambda inst, val: setattr(inst, "text_size", (val[0] - dp(8), None)))
        content.add_widget(msg)

        btn_box = BoxLayout(size_hint_y=None, height=CONTROL_HEIGHT, spacing=dp(10))
        is_danger = "删除" in title or "清空" in title
        btn_ok = (self._make_danger_button if is_danger else self._make_primary_button)("确定")
        btn_ok.size_hint_y = 1
        btn_cancel = self._make_text_button("取消")
        btn_cancel.size_hint_y = 1
        btn_box.add_widget(btn_ok)
        btn_box.add_widget(btn_cancel)
        content.add_widget(btn_box)

        popup = self._make_popup(title, content, (0.88, 0.42))

        confirmed = {"done": False}

        def do_confirm(btn):
            if confirmed["done"]:
                return
            confirmed["done"] = True
            btn.disabled = True
            popup.dismiss()
            confirm_callback(btn)

        btn_ok.bind(on_press=do_confirm)
        btn_cancel.bind(on_press=popup.dismiss)
        popup.open()


class AccountingApp(App):
    def build(self):
        self.title = "个人记账"
        sm = ScreenManager()
        sm.add_widget(MainScreen())
        return sm


if __name__ == "__main__":
    AccountingApp().run()
